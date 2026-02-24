"""
FastAPI app: login a Salchi Monster Restaurant con Chromium,
credenciales en JSON local y cookies de sesión.
"""
from __future__ import annotations

import asyncio
import json
import logging
import queue
import re
import shutil
import sys
import traceback
import uuid as uuid_mod
from pathlib import Path
from typing import Any

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None  # type: ignore

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("restaurant_scraper")

# Asegurar que el proyecto root esté en el path (funciona aunque uvicorn se ejecute desde otra carpeta)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode

from contextlib import asynccontextmanager

from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from pydantic import BaseModel

from app.config import (
    CREDENTIALS_FILE,
    COOKIES_FILE,
    TOKEN_FILE,
    LOGIN_URL,
    LOGIN_API_URL,
    REPORT_URL,
    LOCALES_API_URL,
    REPORTS_DIR,
    REPORTS_LOCALES_JSON,
    REPORTS_CANALES_DELIVERY_JSON,
    DELIVERY_API_BASE,
    DELIVERYS_CACHE_DIR,
    REPORTS_RESTAURANT_MAPS_DIR,
    REPORTS_DIDI_MAPS_DIR,
    UPLOADS_DIR,
    NO_ENTREGADAS_JSON,
    APELACIONES_JSON,
    HORARIOS_JSON,
    PLANILLAS_DIR,
    LOCALES_CONFIG_JSON,
    NOTIFICATIONS_JSON,
    PREFERENCIAS_JSON,
)
from app.router import router as didi_capture_router, run_didi_sedes_prune_loop

# ── Sistema de notificaciones ─────────────────────────────────────────────────
_notif_event_loop: asyncio.AbstractEventLoop | None = None
_notif_queues: dict[str, list[asyncio.Queue]] = {}  # local -> lista de queues WS

MAX_NOTIFS_PER_SEDE = 100
ADMIN_NOTIF_LOCAL = "ADMIN"  # Clave fija para notificaciones dirigidas al admin


def _resolve_local_name(sede: str) -> str:
    """Si 'sede' es un ID numérico, devuelve el nombre del local correspondiente.
    Si ya es un nombre (o ADMIN) lo devuelve tal cual."""
    if not sede or not sede.strip().isdigit():
        return sede
    locales = _read_json(REPORTS_LOCALES_JSON, [])
    if isinstance(locales, list):
        for loc in locales:
            if str(loc.get("id", "")) == sede.strip():
                return loc.get("name", sede)
    return sede


def _read_notificaciones() -> dict:
    data = _read_json(NOTIFICATIONS_JSON, {"items": []})
    if not isinstance(data, dict) or "items" not in data:
        return {"items": []}
    return data


def _write_notificaciones(data: dict) -> None:
    NOTIFICATIONS_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(NOTIFICATIONS_JSON, data)


def _fmt_monto_notif(val) -> str:
    try:
        return f"${int(round(float(val or 0))):,}".replace(",", ".")
    except Exception:
        return str(val or 0)


def _create_notificacion(
    local: str, tipo: str, titulo: str, mensaje: str,
    route_name: str, extra: dict | None = None,
) -> dict:
    notif: dict = {
        "id": str(uuid_mod.uuid4()),
        "local": local,
        "tipo": tipo,
        "titulo": titulo,
        "mensaje": mensaje,
        "leida": False,
        "fecha": _now().isoformat(),
        "route_name": route_name,
        "data": extra or {},
    }
    data = _read_notificaciones()
    items: list = data.get("items", [])
    items.insert(0, notif)
    # Limitar por sede
    sede_items = [i for i in items if i.get("local") == local]
    if len(sede_items) > MAX_NOTIFS_PER_SEDE:
        remove_ids = {i["id"] for i in sede_items[MAX_NOTIFS_PER_SEDE:]}
        items = [i for i in items if i.get("id") not in remove_ids]
    data["items"] = items
    _write_notificaciones(data)
    # Broadcast WS (thread-safe desde endpoints síncronos)
    if _notif_event_loop is not None:
        for q in list(_notif_queues.get(local, [])):
            try:
                _notif_event_loop.call_soon_threadsafe(q.put_nowait, notif)
            except Exception:
                pass
    return notif


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _notif_event_loop
    _notif_event_loop = asyncio.get_running_loop()
    # Carpeta para mapas orderId -> displayNum del capturador Didi (POST /didi/daily-orders-payload)
    app.state.didi_capture_maps_dir = REPORTS_DIR / "didi_maps"
    # Ya no se usa el reporte Excel automático; datos desde API deliverys cada 5 min
    # Locales desde API cada 10 min; primera carga al arranque
    locales_task = asyncio.create_task(_locales_scheduler_loop())
    asyncio.create_task(_fetch_and_save_locales())  # primera vez sin esperar
    logger.info("Locales scheduler: iniciado (cada 10 min)")
    # Migrar deliverys antiguos (un archivo por local) a estructura por fecha
    try:
        _migrate_old_deliverys_to_per_date()
    except Exception as e:
        logger.warning("Migración deliverys: %s", e)
    # Deliverys por local cada 2 min (obtenerDeliverysPorLocalSimple; 5 s entre cada sede)
    deliverys_task = asyncio.create_task(_deliverys_scheduler_loop())
    logger.info("Deliverys scheduler: iniciado (cada 2 min, fecha hoy; 5 s entre sedes)")
    # Login cada 12 h para renovar sesión
    login_refresh_task = asyncio.create_task(_login_refresh_loop())
    logger.info("Login refresh: iniciado (cada 12 h)")
    # Sedes Didi: poda cada 5 s para marcar extensiones desconectadas (>36 s sin heartbeat)
    didi_sedes_task = asyncio.create_task(run_didi_sedes_prune_loop())
    logger.info("Didi sedes prune: iniciado (cada 5 s)")
    # Callback para merge cuando la extensión actualiza didi_restaurant_map (sin temporizador)
    app.state.on_didi_map_updated = _on_didi_map_updated
    yield
    locales_task.cancel()
    deliverys_task.cancel()
    login_refresh_task.cancel()
    didi_sedes_task.cancel()
    try:
        await locales_task
    except asyncio.CancelledError:
        pass
    try:
        await deliverys_task
    except asyncio.CancelledError:
        pass
    try:
        await login_refresh_task
    except asyncio.CancelledError:
        pass
    try:
        await didi_sedes_task
    except asyncio.CancelledError:
        pass


app = FastAPI(
    title="Restaurant Scraper Login",
    description="Login con Chromium a salchimonster.restaurant.pe",
    lifespan=lifespan,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "https://fotopedidos.salchimonster.com",
        "http://fotopedidos.salchimonster.com",
        "null",
    ],
    allow_origin_regex=r"chrome-extension://[a-zA-Z0-9]+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(didi_capture_router, tags=["didi-capture"])

# --- Credenciales ---

def _read_json(path: Path, default: dict | list) -> dict | list:
    if not path.exists():
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return default


def _write_json(path: Path, data: dict | list) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def get_credentials() -> dict[str, Any]:
    data = _read_json(CREDENTIALS_FILE, {})
    if not data:
        raise HTTPException(status_code=503, detail="No hay credenciales. Usa PUT /credentials para configurarlas.")
    return data


def save_credentials(data: dict[str, Any]) -> dict[str, Any]:
    CREDENTIALS_FILE.parent.mkdir(parents=True, exist_ok=True)
    _write_json(CREDENTIALS_FILE, data)
    return data


def get_cookies() -> list[dict]:
    return _read_json(COOKIES_FILE, [])


def save_cookies(cookies: list[dict]) -> None:
    COOKIES_FILE.parent.mkdir(parents=True, exist_ok=True)
    _write_json(COOKIES_FILE, cookies)


def get_token() -> str | None:
    """Devuelve el token guardado (data.token del login) o None."""
    data = _read_json(TOKEN_FILE, {})
    if isinstance(data, dict):
        return data.get("token")
    return None


def save_token(token: str | None) -> None:
    """Guarda el token del login en token.json."""
    if not token:
        return
    TOKEN_FILE.parent.mkdir(parents=True, exist_ok=True)
    from datetime import datetime
    _write_json(TOKEN_FILE, {"token": token, "updated_at": datetime.utcnow().isoformat() + "Z"})


# --- Pydantic models ---

class CredentialsUpdate(BaseModel):
    usuario_nick: str | None = None
    usuario_clave: str | None = None
    usuario_recordar: str | None = None
    local_id: str | None = None
    turno_id: str | None = None
    caja_id: str | None = None
    app: str | None = None


def _clean_server_error_message(raw: str) -> str:
    """Convierte HTML/errores del servidor (ej. Slim PHP) en un mensaje corto y legible."""
    if not raw or not isinstance(raw, str):
        return "Error desconocido del servidor"
    raw = raw.strip()
    # Es una página de error HTML (Slim, etc.)
    if "<html" in raw.lower() or "slim application error" in raw.lower() or "application error" in raw.lower():
        # Extraer el mensaje después de "Message:</strong>" o "Details" (ej. get_object_vars()...)
        m = re.search(r"<strong>\s*Message:\s*</strong>\s*([^<]+)", raw, re.IGNORECASE | re.DOTALL)
        if m:
            detail = m.group(1).strip()
            detail = re.sub(r"\s+", " ", detail)[:200]
            # Mensaje más amigable para el error típico del PHP del restaurante
            if "get_object_vars" in detail and "null given" in detail:
                return (
                    "El servidor del restaurante falló al procesar el login (error interno). "
                    "Comprueba que usuario y clave sean correctos. Si sigue fallando, usa «Probar login por formulario»."
                )
            return f"Error del servidor del restaurante: {detail}"
        m = re.search(r"<strong>\s*Details:\s*</strong>\s*([^<]+)", raw, re.IGNORECASE | re.DOTALL)
        if m:
            detail = m.group(1).strip()
            detail = re.sub(r"\s+", " ", detail)[:200]
            if "get_object_vars" in detail and "null given" in detail:
                return (
                    "El servidor del restaurante falló al procesar el login (error interno). "
                    "Comprueba que usuario y clave sean correctos. Si sigue fallando, usa «Probar login por formulario»."
                )
            return f"Error del servidor del restaurante: {detail}"
        return "Error del servidor del restaurante (página de error). Intenta de nuevo más tarde."
    # Si ya es texto corto sin HTML, detectar el mismo error PHP
    if "get_object_vars" in raw and "null given" in raw:
        return (
            "El servidor del restaurante falló al procesar el login (error interno). "
            "Comprueba que usuario y clave sean correctos. Si sigue fallando, usa «Probar login por formulario»."
        )
    if len(raw) > 400:
        raw = raw[:397] + "..."
    return raw


# --- Endpoints ---

@app.get("/")
def root():
    return {"message": "Restaurant Scraper Login API", "docs": "/docs"}


@app.get("/credentials")
def read_credentials():
    """Devuelve las credenciales guardadas (sin exponer la clave en logs)."""
    cred = get_credentials()
    # Opcional: ofuscar clave en respuesta
    out = cred.copy()
    if out.get("usuario_clave"):
        out["usuario_clave"] = "********"
    return out


# WebSocket y cola para notificar progreso de login (credenciales)
_credentials_ws_clients: list[WebSocket] = []
_login_status_queue: queue.Queue = queue.Queue()


def _login_status_push(step: str, message: str, **extra: Any) -> None:
    """Envía un paso del login a la cola para que se emita por WebSocket (thread-safe)."""
    try:
        _login_status_queue.put_nowait({"step": step, "message": message, **extra})
    except Exception:
        pass


@app.put("/credentials")
def update_credentials(update: CredentialsUpdate):
    """Actualiza las credenciales en credentials.json."""
    current = _read_json(CREDENTIALS_FILE, {})
    payload = update.model_dump(exclude_none=True)
    if not payload:
        raise HTTPException(status_code=400, detail="Envía al menos un campo para actualizar.")
    current.update(payload)
    save_credentials(current)
    return {"message": "Credenciales actualizadas", "credentials": {k: "********" if k == "usuario_clave" else v for k, v in current.items()}}


@app.websocket("/credentials/ws")
async def credentials_status_websocket(websocket: WebSocket):
    """
    WebSocket para recibir en tiempo real el progreso al actualizar credenciales y hacer login.
    Mensajes: step, message y opcionalmente success, credentials.
    """
    await websocket.accept()
    _credentials_ws_clients.append(websocket)
    try:
        while True:
            await asyncio.wait_for(websocket.receive_text(), timeout=300)
    except (WebSocketDisconnect, asyncio.TimeoutError):
        pass
    finally:
        if websocket in _credentials_ws_clients:
            _credentials_ws_clients.remove(websocket)


async def _broadcast_credentials_status(payload: dict) -> None:
    """Envía un mensaje a todos los clientes del WebSocket de credenciales."""
    dead: list[WebSocket] = []
    for ws in list(_credentials_ws_clients):
        try:
            await ws.send_json(payload)
        except Exception:
            dead.append(ws)
    for ws in dead:
        if ws in _credentials_ws_clients:
            _credentials_ws_clients.remove(ws)


@app.post("/credentials/update-and-login")
async def update_credentials_and_login(update: CredentialsUpdate):
    """
    Actualiza las credenciales, intenta login y notifica el progreso por WebSocket (/credentials/ws).
    Si el login es correcto, devuelve las credenciales actuales (clave enmascarada) y la opción de actualizarlas.
    """
    payload = update.model_dump(exclude_none=True)
    if not payload:
        raise HTTPException(status_code=400, detail="Envía al menos un campo (usuario_nick, usuario_clave, etc.).")
    current = _read_json(CREDENTIALS_FILE, {})
    current.update(payload)
    save_credentials(current)
    cred_masked = {k: "********" if k == "usuario_clave" else v for k, v in current.items()}
    await _broadcast_credentials_status({
        "step": "credentials_saved",
        "message": "Credenciales guardadas. Iniciando login...",
    })
    loop = asyncio.get_event_loop()
    future = loop.run_in_executor(_get_executor(), _do_login_sync)
    while not future.done():
        try:
            while True:
                msg = _login_status_queue.get_nowait()
                await _broadcast_credentials_status(msg)
        except queue.Empty:
            pass
        await asyncio.sleep(0.05)
    while True:
        try:
            msg = _login_status_queue.get_nowait()
            await _broadcast_credentials_status(msg)
        except queue.Empty:
            break
    result = future.result()
    success = result.get("success", False) and not result.get("_is_error", False)
    await _broadcast_credentials_status({
        "step": "result",
        "message": result.get("message", "Listo."),
        "success": success,
        "credentials": cred_masked,
    })
    if result.get("_is_error"):
        code = result.get("status_code", 500)
        return JSONResponse(status_code=code, content={
            "message": result.get("message", "Error en login"),
            "success": False,
            "credentials": cred_masked,
        })
    return {
        "message": result.get("message", "Login realizado"),
        "success": True,
        "credentials": cred_masked,
    }


def _value_for_select(cred_value: str) -> str:
    """En el HTML los options tienen value 'string:1'; la API acepta '1'. Intentamos ambos."""
    if not cred_value:
        return cred_value
    return f"string:{cred_value}"


def _do_login_form_sync() -> dict:
    """Login rellenando el formulario de la página (como un usuario). Usa el HTML del login."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return {
            "_is_error": True,
            "status_code": 503,
            "success": False,
            "message": "Playwright no instalado.",
            "saved_cookies": 0,
        }

    cred = get_credentials()
    logger.info("Login (form): inicio - usuario=%s", cred.get("usuario_nick", "?"))

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                ignore_https_errors=True,
            )
            saved_cookies = get_cookies()
            if saved_cookies:
                context.add_cookies(saved_cookies)

            page = context.new_page()
            page.goto(LOGIN_URL, wait_until="networkidle", timeout=30000)

            # Esperar al formulario (Angular)
            page.wait_for_selector('select[name="local_id"]', state="visible", timeout=15000)
            logger.info("Login (form): formulario visible")

            local_val = _value_for_select(cred.get("local_id", "1"))
            caja_val = _value_for_select(cred.get("caja_id", "29"))
            turno_val = _value_for_select(cred.get("turno_id", "1"))

            # Orden: local (dispara carga de caja/turno), luego caja, turno, usuario, clave
            page.select_option('select[name="local_id"]', value=local_val)
            page.wait_for_timeout(1500)  # fnCambiarLocal() carga cajas/turnos

            try:
                page.select_option('select[name="caja_id"]', value=caja_val)
            except Exception as e:
                logger.warning("Login (form): caja_id %s no encontrado, intentando por índice: %s", caja_val, e)
                page.select_option('select[name="caja_id"]', index=1)
            try:
                page.select_option('select[name="turno_id"]', value=turno_val)
            except Exception as e:
                logger.warning("Login (form): turno_id %s no encontrado: %s", turno_val, e)
                page.select_option('select[name="turno_id"]', index=1)

            page.fill('input[name="usuario_nick"]', cred.get("usuario_nick", ""))
            page.fill('input[name="usuario_clave"]', cred.get("usuario_clave", ""))

            # Esperar a que el botón esté habilitado (vm.cargaDataCaja && vm.cargaDataTurno)
            submit = page.locator('button[type="submit"]')
            try:
                submit.wait_for(state="visible", timeout=5000)
                page.wait_for_selector('button[type="submit"]:not([disabled])', timeout=10000)
            except Exception as e:
                logger.warning("Login (form): botón no se habilitó (caja/turno?): %s", e)
            page.wait_for_timeout(500)

            # Esperar respuesta del login al hacer clic (Angular llama al API)
            login_resp_status = None
            token_from_form = None
            try:
                with page.expect_response(lambda r: "usuario/login" in r.url) as resp_info:
                    submit.click()
                    logger.info("Login (form): clic en Iniciar sesión")
                resp = resp_info.value
                login_resp_status = resp.status
                logger.info("Login (form): respuesta login API status=%s", login_resp_status)
                if resp.status == 200:
                    try:
                        body_form = resp.json()
                        if isinstance(body_form, dict) and body_form.get("data") and body_form.get("data").get("token"):
                            token_from_form = body_form["data"]["token"]
                            save_token(token_from_form)
                            logger.info("Login (form): token guardado")
                    except Exception:
                        pass
            except Exception as e:
                logger.warning("Login (form): no se capturó respuesta del API: %s", e)

            # Dar tiempo a redirección / actualización de la vista
            page.wait_for_timeout(5000)
            try:
                page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass

            current_url = page.url
            cookies = context.cookies()
            save_cookies(cookies)

            # Si sigue en login, intentar leer mensaje de error de la página (toast, .has-error, etc.)
            error_hint = ""
            if "#!/login" in current_url:
                try:
                    toast = page.locator(".toast-message, .toast-error, [class*='error']").first
                    if toast.count() > 0 and toast.is_visible():
                        error_hint = " " + (toast.text_content() or "")[:150]
                except Exception:
                    pass
                if not error_hint:
                    try:
                        err_el = page.locator(".has-error .help-block, .alert-danger").first
                        if err_el.count() > 0 and err_el.is_visible():
                            error_hint = " " + (err_el.text_content() or "")[:150]
                    except Exception:
                        pass

            browser.close()

            # Éxito si ya no estamos en la ruta de login
            if "#!/login" in current_url:
                msg = "El formulario se envió pero la página sigue en login (revisa usuario/clave o captcha)."
                if login_resp_status == 500:
                    msg = "El servidor del restaurante respondió 500 al login (mismo error que por API)." + error_hint
                elif error_hint:
                    msg = msg.rstrip(".") + "." + error_hint
                return {
                    "_is_error": True,
                    "status_code": 401,
                    "success": False,
                    "message": msg.strip(),
                    "saved_cookies": len(cookies),
                }

            logger.info("Login (form): éxito, cookies guardadas (%s)", len(cookies))
            return {
                "_is_error": False,
                "success": True,
                "message": "Login realizado (formulario).",
                "token": token_from_form or get_token(),
                "saved_cookies": len(cookies),
            }

    except Exception as e:
        logger.exception("Login (form): excepción %s", e)
        return {
            "_is_error": True,
            "status_code": 500,
            "success": False,
            "message": f"{type(e).__name__}: {e}",
            "saved_cookies": 0,
            "detail": traceback.format_exc(),
        }


def _do_login_sync() -> dict:
    """Ejecuta el login con Playwright síncrono (en proceso separado en Windows para evitar NotImplementedError)."""
    _login_status_push("start", "Iniciando login...")
    logger.info("Login: inicio")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        logger.error("Playwright no instalado: %s", e)
        _login_status_push("error", "Playwright no instalado.", success=False)
        return {
            "_is_error": True,
            "status_code": 503,
            "success": False,
            "message": "Playwright no instalado. Ejecuta: pip install playwright && playwright install chromium",
            "saved_cookies": 0,
        }

    cred = get_credentials()
    _login_status_push("credentials_loaded", "Credenciales cargadas. Lanzando navegador...")
    logger.info("Login: credenciales cargadas (usuario=%s)", cred.get("usuario_nick", "?"))
    form_data = {
        "usuario_nick": cred.get("usuario_nick", ""),
        "usuario_clave": cred.get("usuario_clave", ""),
        "usuario_recordar": cred.get("usuario_recordar", "1"),
        "local_id": cred.get("local_id", "1"),
        "turno_id": cred.get("turno_id", "1"),
        "caja_id": cred.get("caja_id", "29"),
        "app": cred.get("app", "Web"),
    }

    try:
        with sync_playwright() as p:
            _login_status_push("browser_start", "Lanzando Chromium...")
            logger.info("Login: lanzando Chromium")
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                ignore_https_errors=True,
            )

            saved_cookies = get_cookies()
            if saved_cookies:
                context.add_cookies(saved_cookies)

            page = context.new_page()
            _login_status_push("navigating", "Navegando a la página de login...")
            logger.info("Login: navegando a %s", LOGIN_URL)
            page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)

            _login_status_push("sending_login", "Enviando credenciales al servidor...")
            logger.info("Login: POST a %s", LOGIN_API_URL)
            response = page.request.post(
                LOGIN_API_URL,
                form=form_data,
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "Accept": "application/json",
                },
            )

            status = response.status
            logger.info("Login: respuesta HTTP %s", status)
            body = None
            try:
                body = response.json()
                logger.info("Login: body JSON keys=%s", list(body.keys())[:15] if isinstance(body, dict) else "n/a")
            except Exception as e:
                logger.warning("Login: no se pudo parsear JSON: %s", e)
                try:
                    raw = response.text()
                    if raw and raw.strip():
                        body = {"_raw": raw}
                        logger.info("Login: body raw (primeros 200 chars): %s", raw[:200])
                except Exception as e2:
                    logger.warning("Login: no se pudo leer body: %s", e2)

            cookies = context.cookies()
            save_cookies(cookies)
            _login_status_push("saving_session", "Guardando sesión y cookies...")
            logger.info("Login: cookies guardadas (%s)", len(cookies))
            browser.close()

    except Exception as e:
        logger.exception("Login: excepción en Playwright/request: %s", e)
        _login_status_push("error", f"Error: {e}", success=False)
        return {
            "_is_error": True,
            "status_code": 500,
            "success": False,
            "message": f"{type(e).__name__}: {e}",
            "saved_cookies": 0,
            "detail": traceback.format_exc(),
        }

    def _msg(b: dict | None, default: str = "Error desconocido") -> str:
        if not b or not isinstance(b, dict):
            return default
        for key in ("mensajes", "message", "detail", "error"):
            val = b.get(key)
            if isinstance(val, list) and val:
                return _clean_server_error_message(str(val[0]))
            if isinstance(val, str) and val:
                return _clean_server_error_message(val)
        if b.get("_raw"):
            return _clean_server_error_message(str(b["_raw"])[:2000])
        return default

    # Detectar respuesta HTML de error (Slim/PHP) aunque venga con 200
    raw_body = (body or {}).get("_raw", "") if isinstance(body, dict) else ""
    is_html_error = (
        isinstance(raw_body, str)
        and ("<html" in raw_body.lower() or "slim application error" in raw_body.lower())
    )

    if status != 200 or is_html_error:
        msg = _msg(body, f"Login falló (HTTP {status})" if status != 200 else "Error del servidor del restaurante")
        if is_html_error:
            msg = _clean_server_error_message(raw_body)
        _login_status_push("error", msg, success=False)
        logger.warning("Login: falló HTTP %s - %s", status, msg[:200])
        return {
            "_is_error": True,
            "status_code": status if status != 200 else 502,
            "success": False,
            "message": msg,
            "saved_cookies": len(cookies),
        }

    mensaje = "Login realizado"
    if body and isinstance(body, dict) and body.get("mensajes"):
        mensaje = body["mensajes"][0] if body["mensajes"] else mensaje
    else:
        mensaje = _msg(body, mensaje)

    # Extraer y guardar token (data.token) del JSON de login
    token = None
    if isinstance(body, dict) and body.get("data"):
        token = body["data"].get("token")
    if token:
        save_token(token)
        logger.info("Login: token guardado")

    _login_status_push("done", mensaje, success=True)
    logger.info("Login: éxito - %s", mensaje)
    return {
        "_is_error": False,
        "success": True,
        "message": mensaje,
        "tipo": body.get("tipo") if isinstance(body, dict) else None,
        "token": token,
        "saved_cookies": len(cookies),
        "data_preview": list(body.get("data", {}).keys())[:10] if isinstance(body, dict) and body.get("data") else None,
    }


@app.post("/login")
async def do_login(method: str = Query(default="api")):
    """Ejecuta el login usando las credenciales guardadas. method=form usa Playwright por formulario."""
    cred = _read_json(CREDENTIALS_FILE, {})
    if not cred:
        raise HTTPException(status_code=503, detail="No hay credenciales. Usa PUT /credentials para configurarlas.")
    cred_masked = {k: "********" if k == "usuario_clave" else v for k, v in cred.items()}
    fn = _do_login_form_sync if method == "form" else _do_login_sync
    loop = asyncio.get_event_loop()
    future = loop.run_in_executor(_get_executor(), fn)
    while not future.done():
        try:
            while True:
                msg = _login_status_queue.get_nowait()
                await _broadcast_credentials_status(msg)
        except queue.Empty:
            pass
        await asyncio.sleep(0.05)
    while True:
        try:
            msg = _login_status_queue.get_nowait()
            await _broadcast_credentials_status(msg)
        except queue.Empty:
            break
    result = future.result()
    success = result.get("success", False) and not result.get("_is_error", False)
    if result.get("_is_error"):
        code = result.get("status_code", 500)
        return JSONResponse(status_code=code, content={
            "message": result.get("message", "Error en login"),
            "success": False,
            "credentials": cred_masked,
        })
    return {
        "message": result.get("message", "Login realizado"),
        "success": True,
        "credentials": cred_masked,
    }


@app.get("/cookies")
def read_cookies():
    """Lista si hay cookies guardadas (no expone valores sensibles)."""
    cookies = get_cookies()
    return {
        "count": len(cookies),
        "names": [c.get("name") for c in cookies],
    }


@app.post("/cookies/clear")
def clear_cookies():
    """Borra las cookies guardadas (útil para forzar nuevo login)."""
    save_cookies([])
    return {"message": "Cookies borradas"}


@app.get("/token")
def read_token():
    """Devuelve el token guardado (data.token del último login exitoso)."""
    token = get_token()
    if not token:
        raise HTTPException(status_code=404, detail="No hay token. Haz login primero (POST /login).")
    return {"token": token}


# --- Login refresh (renovar sesión cada 12 h) ---

_LOGIN_REFRESH_INTERVAL_SECONDS = 12 * 3600  # 12 horas: hacer login de nuevo para renovar sesión

_login_executor: Any = None


def _get_executor():
    """En Windows usa ProcessPoolExecutor para evitar NotImplementedError de Playwright con subprocesos en threads."""
    global _login_executor
    if _login_executor is None:
        import concurrent.futures
        if sys.platform == "win32":
            _login_executor = concurrent.futures.ProcessPoolExecutor(max_workers=2)
        else:
            _login_executor = concurrent.futures.ThreadPoolExecutor(max_workers=2)
    return _login_executor


async def _login_refresh_loop() -> None:
    """Cada 12 horas ejecuta login de nuevo para renovar la sesión."""
    import asyncio
    while True:
        await asyncio.sleep(_LOGIN_REFRESH_INTERVAL_SECONDS)
        logger.info("Login refresh: ejecutando login (cada 12 h)")
        try:
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(_get_executor(), _do_login_sync)
            if result.get("_is_error"):
                logger.warning("Login refresh: falló - %s", result.get("message", result.get("detail", "error desconocido")))
            else:
                logger.info("Login refresh: OK, token renovado")
        except Exception as e:
            logger.warning("Login refresh: excepción - %s", e)


# --- Informe de ventas (Excel) ---

def _sanitize_path(name: str) -> str:
    """Nombre seguro para carpeta/archivo: sin caracteres inválidos."""
    if not name or not isinstance(name, str):
        return "sin_nombre"
    s = name.strip()
    for c in '\\/:*?"<>|':
        s = s.replace(c, "_")
    return s or "sin_nombre"


def _parse_fecha_to_date_str(fecha: Any) -> str | None:
    """Convierte Fecha (ej. '10-02-2026' o ISO) a 'YYYY-MM-DD'."""
    if not fecha:
        return None
    if hasattr(fecha, "strftime"):
        return fecha.strftime("%Y-%m-%d")
    s = str(fecha).strip()
    if not s:
        return None
    # DD-MM-YYYY
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# Columnas a extraer del Excel (solo filas con fecha). Nombres posibles por columna.
_REPORT_JSON_COLUMNS = [
    ("Fecha", ["Fecha", "fecha"]),
    ("Hora", ["Hora", "hora"]),
    ("Cliente", ["Cliente", "cliente"]),
    ("Local", ["Local", "local"]),
    ("Monto pagado", ["Monto pagado", "Monto pagado", "monto pagado"]),
    ("Canal de delivery", ["Canal de delivery", "canal de delivery"]),
    ("Codigo integracion", ["Codigo integracion", "Codigo integración delivery", "Codigo integración", "codigo integracion"]),
]


def _excel_row_to_json_row(row_values: list, col_indices: dict) -> dict:
    out = {}
    for key, idx in col_indices.items():
        if idx is None:
            out[key] = None
            continue
        val = row_values[idx] if idx < len(row_values) else None
        if key == "Monto pagado" and isinstance(val, (int, float)):
            out[key] = float(val)
        elif val is not None and hasattr(val, "isoformat"):
            try:
                out[key] = val.isoformat()
            except Exception:
                out[key] = str(val)
        else:
            out[key] = str(val).strip() if val is not None and str(val).strip() else None
    return out


def _extract_ventas_json_from_excel(filepath: Path) -> list[dict]:
    """Lee el Excel y devuelve lista de dicts con las columnas indicadas, solo filas con fecha."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    col_indices = {key: None for key, _ in _REPORT_JSON_COLUMNS}
    header_found = False
    out = []

    for row in ws.iter_rows(values_only=True):
        row_list = list(row) if row else []
        row_str = [str(c).strip() if c is not None else "" for c in row_list]

        if not header_found and ("Fecha" in row_str or "fecha" in row_str):
            for key, aliases in _REPORT_JSON_COLUMNS:
                for alias in aliases:
                    if alias in row_str:
                        col_indices[key] = row_str.index(alias)
                        break
            header_found = True
            continue

        if not header_found:
            continue

        idx_fecha = col_indices.get("Fecha")
        if idx_fecha is None or idx_fecha >= len(row_list):
            continue
        fecha_val = row_list[idx_fecha]
        if fecha_val is None or (isinstance(fecha_val, str) and not str(fecha_val).strip()):
            continue
        item = _excel_row_to_json_row(row_list, col_indices)
        item["Fecha"] = fecha_val.isoformat() if hasattr(fecha_val, "isoformat") else str(fecha_val).strip()
        out.append(item)

    wb.close()
    return out


def _save_filas_by_local_day_canal(filas: list[dict]) -> None:
    """
    Guarda las filas en estructura: reports/{Local}/{YYYY-MM-DD}/{Canal delivery}.json
    y actualiza locales.json y canales_delivery.json (listas sin repetir).
    """
    from collections import defaultdict
    groups = defaultdict(list)
    locales_set = set()
    canales_set = set()

    for row in filas:
        local = (row.get("Local") or "").strip() or "Sin local"
        fecha = row.get("Fecha")
        date_str = _parse_fecha_to_date_str(fecha)
        if not date_str:
            continue
        canal = (row.get("Canal de delivery") or "").strip() or "Sin canal"
        key = (local, date_str, canal)
        groups[key].append(row)
        locales_set.add(local)
        canales_set.add(canal)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    for (local, date_str, canal), rows in groups.items():
        dir_local = REPORTS_DIR / _sanitize_path(local) / date_str
        dir_local.mkdir(parents=True, exist_ok=True)
        filepath = dir_local / f"{_sanitize_path(canal)}.json"
        _write_json(filepath, rows)
        logger.info("Report: guardado %s (%s filas)", filepath, len(rows))

    # Actualizar índices: canales sin repetir; locales se fusionan (mantener id+name desde API)
    existing_locales = _locales_list_for_iteration()
    existing_names = {_locale_name(x) for x in existing_locales if _locale_name(x)}
    existing_dicts = []
    for x in existing_locales:
        if isinstance(x, dict) and x.get("name"):
            existing_dicts.append(x)
        elif isinstance(x, str) and x.strip():
            existing_dicts.append({"id": "", "name": x.strip()})
    for name in locales_set:
        if name and name not in existing_names:
            existing_dicts.append({"id": "", "name": name})
            existing_names.add(name)
    existing_dicts.sort(key=lambda x: (x.get("name") or "").lower())
    existing_canales = list(_read_json(REPORTS_CANALES_DELIVERY_JSON, []))
    all_canales = sorted(set(existing_canales) | canales_set)
    REPORTS_LOCALES_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(REPORTS_LOCALES_JSON, existing_dicts)
    _write_json(REPORTS_CANALES_DELIVERY_JSON, all_canales)
    logger.info("Report: índices actualizados (%s locales, %s canales)", len(existing_dicts), len(all_canales))


_REPORT_DEFAULT_PARAMS = {
    "page": "informeventasv3_informeventas",
    "type": "excel",
    "or": "L",
    "caja": "-1",
    "doc": "-1",
    "estado": "1",
    "turno": "-1",
    "ordenarDoc": "0",
    "local": "-1",
    "serie": "-1",
    "numero": "-1",
    "tipoventa": "-1",
    "filtroreservas": "-1",
    "tipoemision": "-1",
    "moneda_id": "1",
    "marca_id": "0",
    "pagina": "1",
    "complementoventa": "0",
    "registros": "20000",
    "soloAlCredito": "0",
    "ruc": "-1",
    "idmonedatc": "-1",
}

# Zona horaria Colombia para "hoy" (en Windows puede requerir pip install tzdata)
_COLOMBIA_TZ = None
if ZoneInfo:
    try:
        _COLOMBIA_TZ = ZoneInfo("America/Bogota")
    except Exception:
        pass


def _now() -> datetime:
    """Ahora en Colombia o UTC (solo UTC si zoneinfo no está disponible)."""
    if _COLOMBIA_TZ:
        return datetime.now(_COLOMBIA_TZ)
    return datetime.utcnow()


def _now_colombia_str() -> str:
    """Fecha y hora actual en Colombia, formato legible para fetched_at: 'YYYY-MM-DD HH:MM:SS (Colombia)'."""
    now = _now()
    if _COLOMBIA_TZ and now.tzinfo:
        return now.strftime("%Y-%m-%d %H:%M:%S") + " (Colombia)"
    # Fallback sin zoneinfo: asumir que _now() es UTC y no etiquetar como Colombia
    return now.strftime("%Y-%m-%d %H:%M:%S") + " (UTC)"


def _get_today_colombia() -> str:
    """Fecha de hoy en Colombia (solo año-mes-día, sin hora)."""
    return _now().strftime("%Y-%m-%d")


def _parse_hhmm(s: str) -> tuple[int, int]:
    """Convierte 'HH:MM' o 'H:MM' en (hour, minute)."""
    s = (s or "").strip()
    if ":" in s:
        parts = s.split(":", 1)
        try:
            return int(parts[0].strip()) % 24, int(parts[1].strip()) % 60
        except ValueError:
            pass
    return 0, 0


def _is_within_opening_hours() -> bool:
    """
    True si la hora actual en Colombia está dentro del horario de apertura (horarios.json).
    Por defecto: 12:30 a 00:00 (medianoche). Fuera de ese horario el restaurante está cerrado.
    """
    data = _read_json(HORARIOS_JSON, {})
    if not isinstance(data, dict):
        return True
    open_at = (data.get("open_at") or "12:30").strip()
    close_at = (data.get("close_at") or "00:00").strip()
    now = _now()
    h, m = now.hour, now.minute
    open_h, open_m = _parse_hhmm(open_at)
    close_h, close_m = _parse_hhmm(close_at)
    now_minutes = h * 60 + m
    open_minutes = open_h * 60 + open_m
    close_minutes = close_h * 60 + close_m
    if close_minutes <= open_minutes:
        # Cierra a medianoche (ej. open 12:30, close 00:00): abierto si now >= open o now < close
        return now_minutes >= open_minutes or now_minutes < close_minutes
    return open_minutes <= now_minutes < close_minutes


async def _run_report_for_date(fecha: str) -> dict:
    """
    Descarga el reporte para una fecha (YYYY-MM-DD), guarda Excel y JSON por local/día/canal.
    Retorna {"success": bool, "error": str|None, "filas": int}.
    """
    import httpx
    token = get_token()
    if not token:
        return {"success": False, "error": "No hay token. Haz login (POST /login).", "filas": 0}
    f1 = f"{fecha} 00:00:00"
    f2 = f"{fecha} 23:59:59"
    name_suffix = datetime.utcnow().strftime("%d.%m.%Y_%H.%M.%S")
    report_name = f"InformeVentas_{name_suffix}"
    params = {**_REPORT_DEFAULT_PARAMS, "name": report_name, "f1": f1, "f2": f2, "token": token}
    url = f"{REPORT_URL}?{urlencode(params)}"
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    cookies_dict = {c["name"]: c["value"] for c in get_cookies() if isinstance(c.get("name"), str) and isinstance(c.get("value"), str)}
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            resp = await client.get(url, cookies=cookies_dict)
    except Exception as e:
        return {"success": False, "error": str(e), "filas": 0}
    if resp.status_code != 200:
        return {"success": False, "error": f"HTTP {resp.status_code}", "filas": 0}
    ct = (resp.headers.get("content-type") or "").lower()
    is_html = "html" in ct or (len(resp.content) < 2000 and resp.content.strip()[:20].lower().startswith(b"<!doctype"))
    if is_html:
        return {"success": False, "error": "Servidor devolvió HTML (token/sesión inválidos).", "filas": 0}
    filename = f"InformeVentas_{fecha}_{fecha}.xlsx"
    filepath = REPORTS_DIR / filename
    filepath.write_bytes(resp.content)
    filas = []
    try:
        filas = _extract_ventas_json_from_excel(filepath)
        _save_filas_by_local_day_canal(filas)
        json_flat = REPORTS_DIR / filename.replace(".xlsx", ".json")
        _write_json(json_flat, filas)
    except Exception as e:
        logger.warning("Report automático: no se pudo generar JSON: %s", e)
    return {"success": True, "error": None, "filas": len(filas)}


# Estado compartido para el programador de reportes y el WebSocket
_report_scheduler_state: dict[str, Any] = {
    "next_run_at": None,  # datetime
    "status": "waiting",  # waiting | calling_report | report_ready
    "last_report_at": None,  # datetime
    "last_error": None,
    "last_filas": 0,
    "interval_seconds": 300,  # 5 min
}
_report_ws_clients: list[WebSocket] = []


async def _report_scheduler_loop() -> None:
    """Cada 5 minutos descarga el reporte del día (fecha hoy Colombia)."""
    state = _report_scheduler_state
    while True:
        now = _now()
        if state["next_run_at"] is None:
            state["next_run_at"] = now
        target = state["next_run_at"]
        if now < target:
            await asyncio.sleep(1)
            continue
        fecha = _get_today_colombia()
        state["status"] = "calling_report"
        state["last_error"] = None
        logger.info("Report automático: descargando reporte del día %s", fecha)
        result = await _run_report_for_date(fecha)
        state["last_report_at"] = _now()
        state["last_filas"] = result.get("filas", 0)
        if result.get("success"):
            state["status"] = "report_ready"
            state["last_error"] = None
            logger.info("Report automático: listo (%s filas)", state["last_filas"])
        else:
            state["status"] = "report_ready"
            state["last_error"] = result.get("error") or "Error desconocido"
            logger.warning("Report automático: falló - %s", state["last_error"])
        state["next_run_at"] = state["last_report_at"].replace(microsecond=0) + timedelta(seconds=state["interval_seconds"])
        await asyncio.sleep(1)


# --- Locales desde API (actualización cada 10 min) ---

_LOCALES_REFRESH_INTERVAL = 600  # 10 minutos


async def _fetch_and_save_locales() -> bool:
    """
    POST a getLocalesPermitidos/0 con Token token="...", parsea data[] y guarda
    en reports/locales.json como [{"id": local_id, "name": local_descripcion}, ...].
    """
    import httpx
    token = get_token()
    if not token:
        logger.debug("Locales API: no hay token, se omite actualización")
        return False
    auth_header = f'Token token="{token}"'
    REPORTS_LOCALES_JSON.parent.mkdir(parents=True, exist_ok=True)
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                LOCALES_API_URL,
                headers={"Authorization": auth_header},
                json={},
            )
    except Exception as e:
        logger.warning("Locales API: error de conexión - %s", e)
        return False
    if resp.status_code != 200:
        logger.warning("Locales API: HTTP %s", resp.status_code)
        return False
    try:
        body = resp.json()
    except Exception:
        logger.warning("Locales API: respuesta no es JSON")
        return False
    if body.get("tipo") == "401":
        logger.debug("Locales API: no autorizado (token inválido o expirado)")
        return False
    data = body.get("data")
    if not isinstance(data, list):
        logger.warning("Locales API: data no es lista")
        return False
    items = []
    for row in data:
        if not isinstance(row, dict):
            continue
        local_id = (row.get("local_id") or "").strip()
        local_descripcion = (row.get("local_descripcion") or "").strip()
        if not local_descripcion:
            continue
        items.append({"id": local_id, "name": local_descripcion})
    items.sort(key=lambda x: (x["name"].lower(), x["id"]))
    _write_json(REPORTS_LOCALES_JSON, items)
    logger.info("Locales API: actualizados %s locales", len(items))
    return True


async def _locales_scheduler_loop() -> None:
    """Cada 10 minutos obtiene la lista de locales desde la API y actualiza locales.json."""
    while True:
        await _fetch_and_save_locales()
        await asyncio.sleep(_LOCALES_REFRESH_INTERVAL)


_DELIVERYS_INTERVAL_SECONDS = 120  # consulta API cada 2 minutos
_DELIVERYS_DELAY_BETWEEN_LOCALS = 5  # segundos entre cada sede para no saturar la API
_DELIVERYS_MAX_PER_LOCAL = 100  # solo los primeros 100 resultados por sede

# Estado compartido para el scheduler de deliverys y el WebSocket /report/ws
_deliverys_scheduler_state: dict[str, Any] = {
    "next_run_at": None,
    "status": "waiting",  # waiting | calling_deliverys | deliverys_ready
    "last_report_at": None,  # datetime (nombre legacy para compat WS)
    "last_error": None,
    "last_filas": 0,  # total deliverys obtenidos en la última pasada
    "interval_seconds": _DELIVERYS_INTERVAL_SECONDS,
}


async def _on_didi_map_updated(date_str: str) -> None:
    """Se llama cuando se actualiza didi_restaurant_map (p. ej. extensión envía daily-orders). Hace merge y notifica."""
    try:
        _build_restaurant_map_for_date(date_str)
        _cross_didi_map_and_update_orders(date_str)
        restaurant_map_path = REPORTS_RESTAURANT_MAPS_DIR / f"restaurant_map_{date_str}.json"
        if restaurant_map_path.exists():
            by_local = _read_json(restaurant_map_path, {})
            if isinstance(by_local, dict):
                for local_id in by_local.keys():
                    payload = {"type": "sede_ready", "local_id": local_id, "fecha": date_str}
                    for ws in list(_report_ws_clients):
                        try:
                            await ws.send_json(payload)
                        except Exception:
                            pass
    except Exception as e:
        logger.debug("Merge al actualizar mapa Didi: %s", e)


async def _deliverys_scheduler_loop() -> None:
    """Cada 2 minutos consulta obtenerDeliverysPorLocalSimple para cada local_id (fecha hoy); espera 5 s entre sedes."""
    import httpx
    state = _deliverys_scheduler_state
    while True:
        await asyncio.sleep(1)
        now = _now()
        if state["next_run_at"] is None:
            state["next_run_at"] = now
        if now < state["next_run_at"]:
            continue
        locales_data = _locales_list_for_iteration()
        local_ids = []
        for item in locales_data:
            lid = _locale_id(item) if isinstance(item, dict) else ""
            if lid:
                local_ids.append(lid)
        if not local_ids:
            state["next_run_at"] = now.replace(microsecond=0) + timedelta(seconds=_DELIVERYS_INTERVAL_SECONDS)
            await asyncio.sleep(1)
            continue
        if not _is_within_opening_hours():
            logger.debug("Deliverys scheduler: fuera de horario de apertura (restaurante cerrado), se omite")
            state["next_run_at"] = now.replace(microsecond=0) + timedelta(seconds=_DELIVERYS_INTERVAL_SECONDS)
            await asyncio.sleep(1)
            continue
        token = get_token()
        cookies_dict = {c["name"]: c["value"] for c in get_cookies() if isinstance(c.get("name"), str) and isinstance(c.get("value"), str)}
        if not token:
            logger.debug("Deliverys scheduler: sin token, se omite (haz login)")
            state["next_run_at"] = now.replace(microsecond=0) + timedelta(seconds=_DELIVERYS_INTERVAL_SECONDS)
            await asyncio.sleep(1)
            continue
        state["status"] = "calling_deliverys"
        state["last_error"] = None
        logger.info("Deliverys scheduler: consultando %s locales (fecha hoy)", len(local_ids))
        total_filas = 0
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                fecha_hoy = _get_today_colombia()
                for i, local_id in enumerate(local_ids):
                    data = await _fetch_deliverys_for_local(client, local_id, cookies_dict, token)
                    _save_deliverys_for_local(local_id, data, consultation_date=fecha_hoy)
                    total_filas += len(data)
                    if i < len(local_ids) - 1:
                        await asyncio.sleep(_DELIVERYS_DELAY_BETWEEN_LOCALS)
            _update_canales_from_deliverys_cache()
            fecha_hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            _build_restaurant_map_for_date(fecha_hoy)
            _cross_didi_map_and_update_orders(fecha_hoy)
            # Notificar a frontend que recargue pedidos (ya con ids Didi reemplazados) para cada sede del día
            restaurant_map_path = REPORTS_RESTAURANT_MAPS_DIR / f"restaurant_map_{fecha_hoy}.json"
            if restaurant_map_path.exists():
                by_local = _read_json(restaurant_map_path, {})
                if isinstance(by_local, dict):
                    for local_id in by_local.keys():
                        payload = {"type": "sede_ready", "local_id": local_id, "fecha": fecha_hoy}
                        for ws in list(_report_ws_clients):
                            try:
                                await ws.send_json(payload)
                            except Exception:
                                pass
            state["status"] = "deliverys_ready"
            state["last_error"] = None
            state["last_filas"] = total_filas
            logger.info("Deliverys scheduler: listo (%s deliverys en total)", total_filas)
        except Exception as e:
            state["status"] = "deliverys_ready"
            state["last_error"] = str(e)
            logger.warning("Deliverys scheduler: error - %s", e)
        state["last_report_at"] = _now()
        state["next_run_at"] = state["last_report_at"].replace(microsecond=0) + timedelta(seconds=_DELIVERYS_INTERVAL_SECONDS)
        await asyncio.sleep(1)


def _locales_list_for_iteration() -> list[dict[str, str] | str]:
    """Devuelve la lista de locales filtrada por blacklist y con renombres aplicados."""
    data = _read_json(REPORTS_LOCALES_JSON, [])
    if not isinstance(data, list):
        return []
    cfg = _read_json(LOCALES_CONFIG_JSON, {})
    blacklist: set[str] = {str(x) for x in (cfg.get("blacklist_ids") or [])}
    rename: dict[str, str] = {str(k): v for k, v in (cfg.get("rename") or {}).items()}
    result = []
    for item in data:
        lid = str(item.get("id", "")) if isinstance(item, dict) else ""
        if lid and lid in blacklist:
            continue
        if lid and lid in rename and isinstance(item, dict):
            item = {**item, "name": rename[lid]}
        result.append(item)
    return result


def _locale_name(item: dict[str, str] | str) -> str:
    """Extrae el nombre del local de un ítem (dict con 'name' o string legacy)."""
    if isinstance(item, dict):
        return (item.get("name") or "").strip()
    return (item or "").strip()


@app.get("/report")
async def report_ventas(
    fecha_inicio: str = Query(..., description="Fecha inicio YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha fin YYYY-MM-DD"),
):
    """
    Descarga el informe de ventas en Excel para el rango de fechas.
    Usa el token y las cookies guardadas. Guarda el archivo en la carpeta reports/ y lo devuelve.
    """
    import httpx

    token = get_token()
    if not token:
        raise HTTPException(status_code=401, detail="No hay token. Haz login primero (POST /login).")

    # Validar y formatear fechas
    try:
        d1 = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d")
        d2 = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Fechas deben ser YYYY-MM-DD (ej: 2026-02-01)")

    f1 = d1.strftime("%Y-%m-%d 00:00:00")
    f2 = d2.strftime("%Y-%m-%d 23:59:59")
    name_suffix = datetime.utcnow().strftime("%d.%m.%Y_%H.%M.%S")
    report_name = f"InformeVentas_{name_suffix}"

    params = {**_REPORT_DEFAULT_PARAMS, "name": report_name, "f1": f1, "f2": f2, "token": token}
    url = f"{REPORT_URL}?{urlencode(params)}"

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    cookies_list = get_cookies()
    cookies_dict = {c["name"]: c["value"] for c in cookies_list if isinstance(c.get("name"), str) and isinstance(c.get("value"), str)}

    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            resp = await client.get(url, cookies=cookies_dict)
    except Exception as e:
        logger.exception("Report: error de conexión %s", e)
        raise HTTPException(status_code=502, detail=f"Error al conectar con el servidor del reporte: {e}")

    if resp.status_code != 200:
        logger.warning("Report: respuesta %s", resp.status_code)
        raise HTTPException(
            status_code=502,
            detail=f"El servidor del reporte respondió {resp.status_code}. ¿Token o sesión expirados? Haz login de nuevo.",
        )

    content_type = (resp.headers.get("content-type") or "").lower()
    is_html = "html" in content_type or (len(resp.content) < 2000 and resp.content.strip()[:20].lower().startswith(b"<!doctype")) or resp.content.strip()[:10].lower().startswith(b"<html")
    if is_html:
        logger.warning("Report: respuesta HTML (posible error o login requerido)")
        raise HTTPException(
            status_code=502,
            detail="El servidor devolvió HTML en lugar de Excel (token/sesión inválidos o error del servidor).",
        )

    filename = f"InformeVentas_{fecha_inicio}_{fecha_fin}.xlsx"
    filepath = REPORTS_DIR / filename
    filepath.write_bytes(resp.content)
    logger.info("Report: guardado %s (%s bytes)", filepath, len(resp.content))

    # Extraer y guardar por carpeta: Local -> día -> archivo por canal delivery + índices
    try:
        filas = _extract_ventas_json_from_excel(filepath)
        _save_filas_by_local_day_canal(filas)
        # También guardar el JSON plano por compatibilidad (mismo nombre que el Excel)
        json_flat = REPORTS_DIR / filename.replace(".xlsx", ".json")
        _write_json(json_flat, filas)
        logger.info("Report: guardado %s (%s filas)", json_flat, len(filas))
    except Exception as e:
        logger.warning("Report: no se pudo generar JSON del Excel: %s", e)

    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/report/locales")
def get_report_locales():
    """Devuelve la lista de locales (sin repetir) registrados en los reportes."""
    return {"locales": _locales_list_for_iteration()}


@app.get("/report/canales-delivery")
def get_report_canales_delivery():
    """Devuelve la lista de canales de delivery (sin repetir) registrados en los reportes."""
    data = _read_json(REPORTS_CANALES_DELIVERY_JSON, [])
    if not isinstance(data, list):
        return {"canales_delivery": []}
    return {"canales_delivery": data}


# --- Órdenes y fotos (frontend) ---

def _sanitize_codigo(codigo: str) -> str:
    """Código seguro para rutas de archivo."""
    s = (codigo or "").strip()
    return "".join(c if c.isalnum() or c in "._-" else "_" for c in s) or "sin_codigo"


# --- Deliverys API (órdenes por local, cada 5 min; reemplaza Excel para listado) ---


def _clean_privacy_name(s: str) -> str:
    """Quita 'privacy protection' y asteriscos de nombres. Deja solo la parte visible."""
    if not s or not isinstance(s, str):
        return ""
    s = s.strip()
    s = re.sub(r"privacy\s+protection\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\*+", "", s)
    return " ".join(s.split()).strip()


def _delivery_row_to_order(row: dict) -> dict:
    """Convierte un ítem de la API obtenerDeliverysPorLocalSimple al formato orden (frontend)."""
    nombres = _clean_privacy_name((row.get("delivery_nombres") or "").strip())
    apellidos = _clean_privacy_name((row.get("delivery_apellidos") or "").strip())
    if apellidos and apellidos != ".":
        cliente = f"{nombres} {apellidos}".strip()
    else:
        cliente = nombres or "—"
    canal_obj = row.get("canaldelivery") or {}
    canal = (canal_obj.get("canaldelivery_descripcion") or row.get("canaldelivery_descripcion") or "").strip() or "—"
    fecha_hora = (row.get("delivery_fecha") or "").strip()
    fecha = fecha_hora[:10] if len(fecha_hora) >= 10 else ""
    hora = fecha_hora[11:19] if len(fecha_hora) >= 19 else (fecha_hora[11:] if len(fecha_hora) > 10 else "")
    importe = (row.get("delivery_importe") or "").strip()
    try:
        monto_val = float(importe) if importe else None
    except (ValueError, TypeError):
        monto_val = None
    codigo_integracion = (
        (row.get("delivery_codigolimadelivery") or row.get("delivery_codigointegracion") or "").strip()
    ) or "—"
    codigo_integracion = _normalize_didi_display_num(codigo_integracion) or codigo_integracion
    if not codigo_integracion:
        codigo_integracion = "—"
    return {
        "Codigo integracion": codigo_integracion,
        "Cliente": cliente,
        "Canal de delivery": canal,
        "Monto pagado": importe if importe else None,
        "Fecha": fecha,
        "Hora": hora,
        "delivery_id": (row.get("delivery_id") or "").strip(),
        "delivery_identificadorunico": (row.get("delivery_identificadorunico") or "").strip(),
        "delivery_orderid_canal": (row.get("delivery_codigolimadelivery_orderid") or "").strip(),
        "delivery_celular": (row.get("delivery_celular") or "").strip(),
    }


def _delivery_auth_header(token: str | None) -> dict[str, str]:
    """Header de autorización para la API de deliverys/reportes: Token token=\"...\"."""
    if not token:
        return {}
    return {"Authorization": f'Token token="{token}"'}


async def _fetch_deliverys_for_local(
    client: "httpx.AsyncClient",
    local_id: str,
    cookies_dict: dict[str, str],
    token: str | None,
) -> list[dict]:
    """Obtiene como máximo los primeros 100 deliverys del local (paginación 50 en 50)."""
    import httpx
    headers = _delivery_auth_header(token)
    page = 1
    page_size = 50
    offset = 0
    all_data: list[dict] = []
    while len(all_data) < _DELIVERYS_MAX_PER_LOCAL:
        url = f"{DELIVERY_API_BASE}/obtenerDeliverysPorLocalSimple/{local_id}/{page}/{page_size}/{offset}"
        try:
            resp = await client.get(url, cookies=cookies_dict, headers=headers)
        except Exception:
            break
        if resp.status_code != 200:
            break
        try:
            body = resp.json()
        except Exception:
            break
        if not isinstance(body, dict):
            # La API a veces devuelve la lista directamente
            if isinstance(body, list):
                data = body
            else:
                data = []
            if data:
                all_data.extend(data)
                if len(all_data) >= _DELIVERYS_MAX_PER_LOCAL:
                    all_data = all_data[:_DELIVERYS_MAX_PER_LOCAL]
            break
        if body.get("tipo") == "401":
            break
        data = body.get("data") if isinstance(body.get("data"), list) else []
        if not data:
            break
        all_data.extend(data)
        if len(all_data) >= _DELIVERYS_MAX_PER_LOCAL:
            all_data = all_data[:_DELIVERYS_MAX_PER_LOCAL]
            break
        if len(data) < page_size:
            break
        offset += page_size
        page += 1
    return all_data


def _migrate_old_deliverys_to_per_date() -> None:
    """Una vez: convierte deliverys/{local_id}.json antiguos a deliverys/{local_id}/{date}.json."""
    if not DELIVERYS_CACHE_DIR.exists():
        return
    for path in DELIVERYS_CACHE_DIR.iterdir():
        if not path.is_file() or path.suffix != ".json":
            continue
        local_id = path.stem
        cached = _read_json(path, {})
        data = cached.get("data") if isinstance(cached.get("data"), list) else []
        if not data:
            path.unlink()
            continue
        from collections import defaultdict
        by_date: dict[str, list[dict]] = defaultdict(list)
        for row in data:
            fecha = (row.get("delivery_fecha") or "").strip()[:10]
            if fecha:
                by_date[fecha].append(row)
        local_dir = DELIVERYS_CACHE_DIR / local_id
        local_dir.mkdir(parents=True, exist_ok=True)
        for date_str, rows in by_date.items():
            filepath = local_dir / f"{date_str}.json"
            out = {"fetched_at": cached.get("fetched_at") or _now_colombia_str(), "data": rows}
            _write_json(filepath, out)
        path.unlink()
        logger.info("Deliverys: migrado %s -> %s fechas", path.name, len(by_date))


def _normalize_didi_display_num(display_num: str) -> str:
    """Quita el # del displayNum Didi para guardar y usar en URLs (ej. #379006 -> 379006)."""
    s = (display_num or "").strip()
    if s.startswith("#"):
        s = s[1:]
    return s


def _looks_like_didi_display_num(cod: str) -> bool:
    """True si el código es displayNum de Didi (#379001 o 379001) y no debe ser sobrescrito por el id largo de la API."""
    s = (cod or "").strip()
    if not s or len(s) > 15:
        return False
    if s.startswith("#") and len(s) > 1 and s[1:].isdigit():
        return True
    if s.isdigit() and 4 <= len(s) <= 10:
        return True
    return False


def _save_deliverys_for_local(local_id: str, data: list[dict], consultation_date: str | None = None) -> None:
    """
    Guarda deliverys en reports/deliverys/{local_id}/{fecha}.json.
    Si consultation_date está definida, el archivo se nombra con esa fecha (día de la consulta).
    Solo agrega o actualiza órdenes con las que vienen del reporte; nunca borra
    las que ya estaban (no se reemplaza el JSON completo por el que llegó).
    Si una fila existente ya tiene delivery_codigolimadelivery como displayNum Didi (#xxx), se preserva al fusionar.
    """
    from collections import defaultdict
    DELIVERYS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    local_dir = DELIVERYS_CACHE_DIR / local_id
    local_dir.mkdir(parents=True, exist_ok=True)
    fetched_at = _now_colombia_str()

    if consultation_date:
        # Un solo archivo con la fecha de consulta; solo órdenes cuya delivery_fecha es ese día
        date_str = (consultation_date or "").strip()[:10]
        if not date_str:
            return
        filepath = local_dir / f"{date_str}.json"
        existing_by_id: dict[str, dict] = {}
        if filepath.exists():
            cached = _read_json(filepath, {})
            if isinstance(cached, list):
                existing_list = cached
            else:
                existing_list = (cached.get("data") or []) if isinstance(cached.get("data"), list) else []
            for i, r in enumerate(existing_list):
                if not isinstance(r, dict):
                    continue
                row_fecha = (r.get("delivery_fecha") or "").strip()[:10]
                if row_fecha != date_str:
                    continue
                did = (r.get("delivery_id") or "").strip()
                key = did if did else f"__existing_{i}"
                existing_by_id[key] = r
        for row in data:
            if not isinstance(row, dict):
                continue
            row_fecha = (row.get("delivery_fecha") or "").strip()[:10]
            if row_fecha != date_str:
                continue
            did = (row.get("delivery_id") or "").strip()
            if did:
                existing_row = existing_by_id.get(did)
                if existing_row and _looks_like_didi_display_num((existing_row.get("delivery_codigolimadelivery") or "").strip()):
                    row = dict(row)
                    row["delivery_codigolimadelivery"] = (existing_row.get("delivery_codigolimadelivery") or "").strip()
                existing_by_id[did] = row
            else:
                existing_by_id[f"__new_{len(existing_by_id)}"] = row
        out = {"fetched_at": fetched_at, "data": list(existing_by_id.values())}
        _write_json(filepath, out)
        logger.debug("Deliverys: fusionados %s ítems para local_id=%s fecha=%s", len(out["data"]), local_id, date_str)
        return

    by_date: dict[str, list[dict]] = defaultdict(list)
    for row in data:
        if not isinstance(row, dict):
            continue
        fecha = (row.get("delivery_fecha") or "").strip()[:10]
        if fecha:
            by_date[fecha].append(row)
    for date_str, new_rows in by_date.items():
        filepath = local_dir / f"{date_str}.json"
        existing_by_id = {}
        if filepath.exists():
            cached = _read_json(filepath, {})
            if isinstance(cached, list):
                existing_list = cached
            else:
                existing_list = (cached.get("data") or []) if isinstance(cached.get("data"), list) else []
            for i, r in enumerate(existing_list):
                if not isinstance(r, dict):
                    continue
                did = (r.get("delivery_id") or "").strip()
                key = did if did else f"__existing_{i}"
                existing_by_id[key] = r
        for r in new_rows:
            did = (r.get("delivery_id") or "").strip()
            if did:
                existing_row = existing_by_id.get(did)
                if existing_row and _looks_like_didi_display_num((existing_row.get("delivery_codigolimadelivery") or "").strip()):
                    r = dict(r)
                    r["delivery_codigolimadelivery"] = (existing_row.get("delivery_codigolimadelivery") or "").strip()
                existing_by_id[did] = r
            else:
                existing_by_id[f"__new_{len(existing_by_id)}"] = r
        out = {"fetched_at": fetched_at, "data": list(existing_by_id.values())}
        _write_json(filepath, out)
        logger.debug("Deliverys: fusionados %s ítems para local_id=%s fecha=%s", len(out["data"]), local_id, date_str)


def _update_canales_from_deliverys_cache() -> None:
    """Actualiza canales_delivery.json con los canales presentes en la cache de deliverys (por local/fecha)."""
    canales_set = set()
    if DELIVERYS_CACHE_DIR.exists():
        for json_file in DELIVERYS_CACHE_DIR.rglob("*.json"):
            if not json_file.is_file():
                continue
            cached = _read_json(json_file, {})
            if isinstance(cached, list):
                data = cached
            else:
                data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                if not isinstance(row, dict):
                    continue
                canal_obj = row.get("canaldelivery") or {}
                desc = (canal_obj.get("canaldelivery_descripcion") or row.get("canaldelivery_descripcion") or "").strip()
                if desc:
                    canales_set.add(desc)
    if canales_set:
        existing = list(_read_json(REPORTS_CANALES_DELIVERY_JSON, []))
        all_canales = sorted(set(existing) | canales_set)
        REPORTS_CANALES_DELIVERY_JSON.parent.mkdir(parents=True, exist_ok=True)
        _write_json(REPORTS_CANALES_DELIVERY_JSON, all_canales)


def _build_restaurant_map_for_date(date_str: str) -> None:
    """Construye restaurant_map_{date}.json con sede (local_id) e id de pedido solo para Didi (canal Didi Food)."""
    date_str = (date_str or "").strip()[:10]
    if not date_str:
        return
    # Por local: lista de delivery_codigolimadelivery que son Didi
    by_local: dict[str, list[str]] = {}
    if not DELIVERYS_CACHE_DIR.exists():
        return
    for local_dir in DELIVERYS_CACHE_DIR.iterdir():
        if not local_dir.is_dir():
            continue
        local_id = local_dir.name
        filepath = local_dir / f"{date_str}.json"
        if not filepath.exists():
            continue
        cached = _read_json(filepath, {})
        data = cached.get("data") if isinstance(cached, dict) and isinstance(cached.get("data"), list) else []
        for row in data:
            if not isinstance(row, dict):
                continue
            canal_obj = row.get("canaldelivery") or {}
            desc = (canal_obj.get("canaldelivery_descripcion") or row.get("canaldelivery_descripcion") or "").strip()
            if desc != "Didi Food":
                continue
            cod = (row.get("delivery_codigolimadelivery") or "").strip()
            if not cod:
                continue
            if local_id not in by_local:
                by_local[local_id] = []
            if cod not in by_local[local_id]:
                by_local[local_id].append(cod)
    REPORTS_RESTAURANT_MAPS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = REPORTS_RESTAURANT_MAPS_DIR / f"restaurant_map_{date_str}.json"
    _write_json(out_path, by_local)
    logger.debug("Restaurant map: %s escrita para %s (%s locales)", out_path.name, date_str, len(by_local))


def _cross_didi_map_and_update_orders(date_str: str) -> None:
    """Cruza restaurant_map con didi_restaurant_map; actualiza filas con delivery_displaynum_didi y unifica fotos en uploads."""
    date_str = (date_str or "").strip()[:10]
    if not date_str:
        return
    restaurant_map_path = REPORTS_RESTAURANT_MAPS_DIR / f"restaurant_map_{date_str}.json"
    didi_map_path = REPORTS_DIDI_MAPS_DIR / f"didi_restaurant_map_{date_str}.json"
    if not restaurant_map_path.exists() or not didi_map_path.exists():
        return
    by_local = _read_json(restaurant_map_path, {})
    if not isinstance(by_local, dict):
        return
    didi_map = _read_json(didi_map_path, {})
    if not isinstance(didi_map, dict):
        return
    # didi_map: orderId (codigo_lima) -> displayNum (ej. "#597026")
    updates_by_file: dict[Path, list[dict]] = {}  # filepath -> list of rows to write back
    for local_id, codigos in by_local.items():
        if not codigos:
            continue
        filepath = DELIVERYS_CACHE_DIR / local_id / f"{date_str}.json"
        if not filepath.exists():
            continue
        cached = _read_json(filepath, {})
        data = cached.get("data") if isinstance(cached, dict) and isinstance(cached.get("data"), list) else []
        if not data:
            continue
        changed = False
        for row in data:
            if not isinstance(row, dict):
                continue
            cod = (row.get("delivery_codigolimadelivery") or "").strip()
            if cod not in didi_map:
                continue
            display_num = didi_map.get(cod)
            if not display_num:
                continue
            if isinstance(display_num, str):
                display_num = display_num.strip()
            display_num = _normalize_didi_display_num(display_num)
            if not display_num:
                continue
            # Primero unificar fotos: copiar uploads/{codigo_lima} -> uploads/{displayNum sin #} para no perder fotos
            src_base = UPLOADS_DIR / _sanitize_codigo(cod)
            if src_base.exists() and src_base.is_dir():
                dst_base = UPLOADS_DIR / _sanitize_codigo(display_num)
                dst_base.mkdir(parents=True, exist_ok=True)
                for sub in ("entrega", "apelacion", "respuestas"):
                    src_sub = src_base / sub
                    if not src_sub.is_dir():
                        continue
                    dst_sub = dst_base / sub
                    dst_sub.mkdir(parents=True, exist_ok=True)
                    for f in src_sub.rglob("*"):
                        if f.is_file():
                            rel = f.relative_to(src_sub)
                            dest_file = dst_sub / rel
                            dest_file.parent.mkdir(parents=True, exist_ok=True)
                            if not dest_file.exists():
                                try:
                                    shutil.copy2(f, dest_file)
                                except OSError as e:
                                    logger.warning("No se pudo copiar foto %s -> %s: %s", f, dest_file, e)
            # Reemplazar el id por el displayNum (así Codigo integracion y fotos usan el mismo valor)
            row["delivery_codigolimadelivery"] = display_num
            changed = True
        if changed:
            updates_by_file[filepath] = data
            cached["data"] = data
            _write_json(filepath, cached)
    if updates_by_file:
        logger.debug("Didi map cruzado para %s: actualizados %s archivos deliverys", date_str, len(updates_by_file))


def _locale_id(item: dict[str, str] | str) -> str:
    """Extrae el id del local de un ítem (dict con 'id' o vacío)."""
    if isinstance(item, dict):
        return (item.get("id") or "").strip()
    return ""


def _get_local_id_by_name(local_name: str) -> str | None:
    """Devuelve el local_id para un nombre de local (desde locales.json)."""
    for item in _locales_list_for_iteration():
        if _locale_name(item) == (local_name or "").strip():
            lid = _locale_id(item) if isinstance(item, dict) else ""
            return lid if lid else None
    return None


def _get_orders_for_local_date(local: str, fecha: str) -> list[dict]:
    """Órdenes para un local y fecha desde reports/deliverys/{local_id}/{YYYY-MM-DD}.json."""
    local_id = _get_local_id_by_name(local)
    if not local_id:
        return []
    date_str = fecha.strip()[:10]
    filepath = DELIVERYS_CACHE_DIR / local_id / f"{date_str}.json"
    if not filepath.exists():
        return []
    cached = _read_json(filepath, {})
    data = cached.get("data") if isinstance(cached.get("data"), list) else []
    return [_delivery_row_to_order(row) for row in data]


def _find_order_by_codigo(codigo: str) -> dict | None:
    """Busca una orden por código de integración o identificador único en deliverys/{local_id}/{fecha}.json."""
    cod = (codigo or "").strip().lstrip("#")
    if not cod:
        return None
    if not DELIVERYS_CACHE_DIR.exists():
        return None
    for local_dir in DELIVERYS_CACHE_DIR.iterdir():
        if not local_dir.is_dir():
            continue
        for json_file in local_dir.glob("*.json"):
            cached = _read_json(json_file, {})
            data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                cod_lima = (row.get("delivery_codigolimadelivery") or row.get("delivery_codigointegracion") or "").strip()
                identificador = (row.get("delivery_identificadorunico") or "").strip()
                orderid_canal = (row.get("delivery_codigolimadelivery_orderid") or "").strip()
                cod_lima_norm = _normalize_didi_display_num(cod_lima) or cod_lima
                if cod_lima == cod or cod_lima_norm == cod or identificador == cod or orderid_canal == cod:
                    return _delivery_row_to_order(row)
    return None


def _find_orders_by_codigo_partial(query: str, limit: int = 30) -> list[dict]:
    """Busca órdenes cuyo código de integración, identificador único u orden canal contengan la cadena (substring, insensible a mayúsculas/minúsculas)."""
    q = (query or "").strip().lstrip("#").lower()
    if not q:
        return []
    if not DELIVERYS_CACHE_DIR.exists():
        return []
    results: list[dict] = []
    seen_ids: set[str] = set()
    for local_dir in DELIVERYS_CACHE_DIR.iterdir():
        if not local_dir.is_dir():
            continue
        for json_file in local_dir.glob("*.json"):
            cached = _read_json(json_file, {})
            data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                cod_lima = (row.get("delivery_codigolimadelivery") or row.get("delivery_codigointegracion") or "").strip()
                identificador = (row.get("delivery_identificadorunico") or "").strip()
                orderid_canal = (row.get("delivery_codigolimadelivery_orderid") or "").strip()
                cod_lima_norm = _normalize_didi_display_num(cod_lima) or cod_lima
                searchable = f"{cod_lima} {cod_lima_norm} {identificador} {orderid_canal}".lower()
                if q in searchable:
                    uid = cod_lima or identificador or orderid_canal
                    if uid and uid in seen_ids:
                        continue
                    if uid:
                        seen_ids.add(uid)
                    results.append(_delivery_row_to_order(row))
                    if len(results) >= limit:
                        return results
    return results


def _get_fotos_for_codigo(codigo: str) -> dict:
    """Devuelve { entrega: [urls], apelacion: { canal: [urls] } } para un código (carpeta uploads)."""
    cod = (codigo or "").strip().lstrip("#")
    base = _uploads_base_for_codigo(codigo)
    out = {"entrega": [], "apelacion": {}}
    if not base.exists():
        return out
    entrega_folder = base / "entrega"
    if entrega_folder.is_dir():
        out["entrega"] = [f"/api/orders/{cod}/fotos/entrega/{f.name}" for f in entrega_folder.iterdir() if f.is_file()]
    apelacion_dir = base / "apelacion"
    if apelacion_dir.is_dir():
        for canal_dir in apelacion_dir.iterdir():
            if canal_dir.is_dir():
                out["apelacion"][canal_dir.name] = [f"/api/orders/{cod}/fotos/apelacion/{canal_dir.name}/{f.name}" for f in canal_dir.iterdir() if f.is_file()]
    return out


def _foto_codigo_candidates(order: dict) -> list[str]:
    """Códigos a probar para buscar fotos de una orden (referencia canónica, orderId canal y alternativas)."""
    cod_integ = (order.get("Codigo integracion") or "").strip()
    identificador = (order.get("delivery_identificadorunico") or "").strip()
    orderid_canal = (order.get("delivery_orderid_canal") or "").strip()
    seen = set()
    out = []
    for c in (cod_integ, orderid_canal, identificador):
        if c and c != "—" and c not in seen:
            seen.add(c)
            out.append(c)
    return out


def _get_fotos_for_order(order: dict) -> dict:
    """Fotos de la orden probando todas las referencias (codigo integración e identificador unico). Así funcionan las que ya tienen foto en otra carpeta."""
    merged = {"entrega": [], "apelacion": {}}
    for cod in _foto_codigo_candidates(order):
        fotos = _get_fotos_for_codigo(cod)
        for url in fotos.get("entrega", []):
            if url not in merged["entrega"]:
                merged["entrega"].append(url)
        for canal, urls in (fotos.get("apelacion") or {}).items():
            for url in urls:
                if url not in merged["apelacion"].setdefault(canal, []):
                    merged["apelacion"][canal].append(url)
    return merged


def _organize_foto_refs() -> dict:
    """
    Renombra/organiza carpetas en uploads/ para que las fotos queden bajo la referencia canónica
    (Codigo integracion). Si una carpeta existe por identificador unico y otra por codigo integración,
    mueve el contenido a la carpeta del código de integración y elimina la carpeta duplicada.
    """
    if not DELIVERYS_CACHE_DIR.exists():
        return {"moved": [], "errors": [], "message": "No hay cache de deliverys"}
    # Construir mapa: carpeta_alternativa -> carpeta_canonica (sanitized)
    alt_to_canon: dict[str, str] = {}
    for local_dir in DELIVERYS_CACHE_DIR.iterdir():
        if not local_dir.is_dir():
            continue
        for json_file in local_dir.glob("*.json"):
            cached = _read_json(json_file, {})
            data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                o = _delivery_row_to_order(row)
                can = _sanitize_codigo((o.get("Codigo integracion") or "").strip())
                ident = (o.get("delivery_identificadorunico") or "").strip()
                alt = _sanitize_codigo(ident) if ident and ident != "—" else ""
                if alt and can and alt != can:
                    alt_to_canon[alt] = can
    moved = []
    errors = []
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    for alt_folder, can_folder in alt_to_canon.items():
        src = UPLOADS_DIR / alt_folder
        if not src.is_dir():
            continue
        dst = UPLOADS_DIR / can_folder
        try:
            if not dst.exists():
                shutil.copytree(src, dst)
                shutil.rmtree(src)
                moved.append({"from": alt_folder, "to": can_folder})
            else:
                for sub in ("entrega", "apelacion"):
                    sub_src = src / sub
                    if not sub_src.is_dir():
                        continue
                    sub_dst = dst / sub
                    sub_dst.mkdir(parents=True, exist_ok=True)
                    if sub == "apelacion":
                        for canal_dir in sub_src.iterdir():
                            if canal_dir.is_dir():
                                (sub_dst / canal_dir.name).mkdir(parents=True, exist_ok=True)
                                for f in canal_dir.iterdir():
                                    if f.is_file():
                                        shutil.copy2(f, sub_dst / canal_dir.name / f.name)
                    else:
                        for f in sub_src.iterdir():
                            if f.is_file():
                                shutil.copy2(f, sub_dst / f.name)
                shutil.rmtree(src)
                moved.append({"from": alt_folder, "to": can_folder, "merged": True})
        except Exception as e:
            errors.append({"folder": alt_folder, "error": str(e)})
    return {"moved": moved, "errors": errors, "message": f"Organizadas {len(moved)} carpetas" if moved else "Nada que organizar"}


def _get_no_entregadas_set() -> set[str]:
    """Lee la lista de delivery_id marcados como no entregada."""
    data = _read_json(NO_ENTREGADAS_JSON, [])
    if not isinstance(data, list):
        return set()
    return {str(x).strip() for x in data if x}


def _mark_no_entregada(delivery_id: str) -> None:
    """Marca una orden como no entregada (por delivery_id)."""
    did = (delivery_id or "").strip()
    if not did:
        return
    ids_set = _get_no_entregadas_set()
    ids_set.add(did)
    NO_ENTREGADAS_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(NO_ENTREGADAS_JSON, sorted(ids_set))
    logger.info("No entregada: marcado delivery_id=%s", did)


def _unmark_no_entregada(delivery_id: str) -> None:
    """Quita la marca de no entregada cuando se sube foto de entrega."""
    did = (delivery_id or "").strip()
    if not did:
        return
    ids_set = _get_no_entregadas_set()
    if did not in ids_set:
        return
    ids_set.discard(did)
    NO_ENTREGADAS_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(NO_ENTREGADAS_JSON, sorted(ids_set))
    logger.info("No entregada: quitada marca delivery_id=%s (foto de entrega subida)", did)


def _order_has_entrega_photo(codigo: str) -> bool:
    """True si la orden tiene al menos una foto en uploads (misma lógica que servir fotos: base + fallback sin #)."""
    if not (codigo or "").strip() or (codigo or "").strip() == "—":
        return False
    base = _uploads_base_for_codigo(codigo) / "entrega"
    return base.is_dir() and any(base.iterdir())


def _order_has_entrega_photo_from_order(order: dict) -> bool:
    """True si la orden tiene foto de entrega en alguna de sus referencias (codigo integración o identificador unico)."""
    for cod in _foto_codigo_candidates(order):
        if _order_has_entrega_photo(cod):
            return True
    return False


# --- Apelaciones (marcar para apelar / apelar / reporte) ---

def _read_apelaciones() -> dict:
    data = _read_json(APELACIONES_JSON, {"items": []})
    if not isinstance(data, dict) or "items" not in data:
        return {"items": []}
    if not isinstance(data["items"], list):
        return {"items": []}
    return data


def _write_apelaciones(data: dict) -> None:
    APELACIONES_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(APELACIONES_JSON, data)


def _get_apelacion_by_codigo(codigo: str) -> dict | None:
    data = _read_apelaciones()
    cod = (codigo or "").strip()
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() == cod:
            return item
    return None


def _total_reembolsado(item: dict) -> float:
    """Suma de todos los reembolsos (incrementales). Compat con legacy: monto_reembolsado único."""
    reembolsos = item.get("reembolsos")
    if isinstance(reembolsos, list):
        return sum(float(r.get("monto") or 0) for r in reembolsos if isinstance(r, dict))
    if item.get("reembolsado") and item.get("monto_reembolsado") is not None:
        return float(item.get("monto_reembolsado") or 0)
    return 0.0


def _total_descuentos_sede(item: dict) -> float:
    """Suma de descuentos ejecutados (hechos efectivos) a la sede. Compat con legacy: sin campo ejecutado = True."""
    descuentos = item.get("descuentos")
    if isinstance(descuentos, list):
        return sum(
            float(d.get("monto") or 0)
            for d in descuentos
            if isinstance(d, dict) and d.get("ejecutado", True)
        )
    if item.get("descuento_confirmado") and _calcular_perdida_antes_descuento(item) > 0:
        return _calcular_perdida_antes_descuento(item)
    return 0.0


def _total_descuentos_programados(item: dict) -> float:
    """Suma de todos los descuentos programados (ejecutados + pendientes de ejecutar)."""
    descuentos = item.get("descuentos")
    if isinstance(descuentos, list):
        return sum(float(d.get("monto") or 0) for d in descuentos if isinstance(d, dict))
    if item.get("descuento_confirmado") and _calcular_perdida_antes_descuento(item) > 0:
        return _calcular_perdida_antes_descuento(item)
    return 0.0


def _monto_empresa_asume(item: dict) -> float:
    """Monto asumido por la empresa (no se descuenta a nadie)."""
    return float(item.get("monto_empresa_asume") or 0)


def _calcular_perdida_antes_descuento(item: dict) -> float:
    """Pérdida = monto_descontado - total reembolsado por el canal (antes de descontar a la sede)."""
    descontado = float(item.get("monto_descontado") or 0)
    devuelto = _total_reembolsado(item)
    return max(0.0, descontado - devuelto)


def _order_has_respuesta_foto(codigo: str) -> bool:
    """True si la orden tiene al menos una foto en respuestas (respuesta del canal). Misma base que servir fotos."""
    if not (codigo or "").strip():
        return False
    base = _uploads_base_for_codigo(codigo) / "respuestas"
    return base.is_dir() and any(f for f in base.iterdir() if f.is_file())


def _get_orders_for_local_date_range(local: str, fecha_desde: str, fecha_hasta: str) -> list[dict]:
    """Órdenes para un local en el rango de fechas (inclusive). Cada orden tiene Fecha del día."""
    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    if not desde or not hasta:
        return []
    if desde > hasta:
        desde, hasta = hasta, desde  # normalizar orden
    from datetime import datetime, timedelta
    try:
        d1 = datetime.strptime(desde, "%Y-%m-%d").date()
        d2 = datetime.strptime(hasta, "%Y-%m-%d").date()
    except ValueError:
        return []
    orders = []
    current = d1
    while current <= d2:
        date_str = current.strftime("%Y-%m-%d")
        day_orders = _get_orders_for_local_date(local, date_str)
        for o in day_orders:
            o["Fecha"] = date_str  # asegurar fecha del día
            orders.append(o)
        current += timedelta(days=1)
    return orders


@app.get("/api/orders")
def api_get_orders(
    local: str | None = Query(None, description="Nombre de un solo local/sede (opcional si se usa locales)"),
    locales: str | None = Query(None, description="Nombres de locales separados por coma; procesa todas las sedes en una sola petición"),
    fecha: str | None = Query(None, description="Fecha YYYY-MM-DD (un solo día); opcional si se usan fecha_desde y fecha_hasta"),
    fecha_desde: str = Query("", description="Inicio rango YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="Fin rango YYYY-MM-DD"),
    exclude_marcadas_apelacion: bool = Query(False, description="Excluir órdenes ya marcadas para apelación"),
):
    """Lista órdenes. Con 'locales' se procesan todas las sedes en el backend (recomendado para marcar apelación)."""
    # Determinar lista de sedes: locales (varias) o local (una)
    sede_names: list[str] = []
    if (locales or "").strip():
        sede_names = [s.strip() for s in locales.strip().split(",") if s.strip()]
    if not sede_names and (local or "").strip():
        sede_names = [local.strip()]
    if not sede_names:
        raise HTTPException(status_code=400, detail="Indica 'local' o 'locales' (nombres separados por coma)")

    # Prioridad: rango (fecha_desde + fecha_hasta) sobre un solo día (fecha)
    use_range = (fecha_desde or "").strip() and (fecha_hasta or "").strip()
    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    f_single = (fecha or "").strip()[:10]

    orders: list[dict] = []
    for sede in sede_names:
        if use_range:
            site_orders = _get_orders_for_local_date_range(sede, desde, hasta)
        else:
            if not f_single:
                continue
            site_orders = _get_orders_for_local_date(sede, f_single)
        for o in site_orders:
            o["Local"] = sede
            o["rowKey"] = f"{sede}-{(o.get('Codigo integracion') or '').strip()}-{o.get('Fecha') or ''}"
            orders.append(o)

    if not use_range and not f_single:
        return {"orders": []}

    if exclude_marcadas_apelacion:
        apelaciones = _read_apelaciones()
        codigos_marcados = {(item.get("codigo") or "").strip() for item in apelaciones.get("items", [])}
        orders = [o for o in orders if (o.get("Codigo integracion") or "").strip() not in codigos_marcados]

    no_entregadas = _get_no_entregadas_set()
    for o in orders:
        o["has_entrega_photo"] = _order_has_entrega_photo_from_order(o)
        fotos = _get_fotos_for_order(o)
        o["fotos_entrega"] = fotos.get("entrega", [])
        o["no_entregada"] = (o.get("delivery_id") or "").strip() in no_entregadas
    return {"orders": orders}


@app.get("/api/orders/by-codigo/{codigo:path}")
def api_get_order_by_codigo(codigo: str):
    """Devuelve la orden buscando por código de integración o identificador único; fotos y flags has_entrega_photo, no_entregada. Fotos se buscan en todas las referencias (codigo integración e identificador unico)."""
    order = _find_order_by_codigo(codigo)
    if order:
        fotos = _get_fotos_for_order(order)
        order["has_entrega_photo"] = _order_has_entrega_photo_from_order(order)
        order["no_entregada"] = (order.get("delivery_id") or "").strip() in _get_no_entregadas_set()
    else:
        fotos = _get_fotos_for_codigo(codigo)
    return {"order": order, "fotos": fotos}


@app.get("/api/orders/search")
def api_search_orders(q: str = ""):
    """Búsqueda parcial de órdenes por código de integración o identificador único (substring, case-insensitive). Devuelve hasta 30 coincidencias con flags has_entrega_photo y no_entregada."""
    q = (q or "").strip()
    if not q:
        return {"orders": [], "exact": False}
    no_entregadas = _get_no_entregadas_set()
    # Primero intenta coincidencia exacta
    exact = _find_order_by_codigo(q)
    if exact:
        exact["has_entrega_photo"] = _order_has_entrega_photo_from_order(exact)
        exact["no_entregada"] = (exact.get("delivery_id") or "").strip() in no_entregadas
        fotos = _get_fotos_for_order(exact)
        return {"orders": [exact], "exact": True, "fotos": fotos}
    # Si no hay exacta, búsqueda parcial
    results = _find_orders_by_codigo_partial(q)
    for o in results:
        o["has_entrega_photo"] = _order_has_entrega_photo_from_order(o)
        o["no_entregada"] = (o.get("delivery_id") or "").strip() in no_entregadas
    return {"orders": results, "exact": False}


class NoEntregadaBody(BaseModel):
    delivery_id: str


@app.post("/api/orders/no-entregada")
def api_mark_no_entregada(body: NoEntregadaBody):
    """Marca la orden como no entregada (para seguimiento del cajero)."""
    _mark_no_entregada(body.delivery_id)
    return {"ok": True, "delivery_id": body.delivery_id.strip()}


class MarcarApelacionBody(BaseModel):
    codigo: str
    canal: str
    delivery_id: str
    monto_descontado: float
    local: str = ""
    fecha: str = ""


class ApelarBody(BaseModel):
    codigo: str
    monto_devuelto: float


@app.post("/api/apelaciones/marcar")
def api_marcar_apelacion(body: MarcarApelacionBody):
    """Admin: marca una orden para apelación con el monto que nos descontó el canal."""
    data = _read_apelaciones()
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    items = data.get("items", [])
    local = (body.local or "").strip()
    fecha = (body.fecha or "").strip()
    monto_desc = float(body.monto_descontado) if body.monto_descontado is not None else 0
    canal = (body.canal or "").strip()
    # Actualizar si ya existe
    for item in items:
        if (item.get("codigo") or "").strip() == cod:
            item["canal"] = canal
            item["delivery_id"] = (body.delivery_id or "").strip()
            item["monto_descontado"] = monto_desc
            item["fecha_marcado"] = _now().isoformat()
            item["local"] = local
            item["fecha"] = fecha
            _write_apelaciones(data)
            if local:
                _create_notificacion(
                    local=local, tipo="orden_por_apelar",
                    titulo="Pedido por apelar",
                    mensaje=f"El pedido #{cod} ({canal}) por {_fmt_monto_notif(monto_desc)} está pendiente de apelación.",
                    route_name="apelar",
                    extra={"codigo": cod, "canal": canal, "monto": monto_desc},
                )
            return {"ok": True}
    items.append({
        "codigo": cod,
        "canal": canal,
        "delivery_id": (body.delivery_id or "").strip(),
        "monto_descontado": monto_desc,
        "monto_devuelto": None,
        "fecha_marcado": _now().isoformat(),
        "fecha_apelado": None,
        "local": local,
        "fecha": fecha,
    })
    data["items"] = items
    _write_apelaciones(data)
    if local:
        _create_notificacion(
            local=local, tipo="orden_por_apelar",
            titulo="Pedido por apelar",
            mensaje=f"El pedido #{cod} ({canal}) por {_fmt_monto_notif(monto_desc)} está pendiente de apelación.",
            route_name="apelar",
            extra={"codigo": cod, "canal": canal, "monto": monto_desc},
        )
    return {"ok": True}


@app.get("/api/apelaciones/pendientes")
def api_apelaciones_pendientes(
    local: str = Query(..., description="Nombre del local/sede"),
    fecha: str = Query("", description="Fecha YYYY-MM-DD (un día; opcional si usas fecha_desde/fecha_hasta)"),
    fecha_desde: str = Query("", description="Inicio rango YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="Fin rango YYYY-MM-DD"),
):
    """Órdenes marcadas para apelación que aún no tienen respuesta (foto + monto_devuelto). Para la vista Apelar del user."""
    config = _get_app_config()
    dias_para_apelar = int(config.get("dias_para_apelar") or 5)
    apelaciones = _read_apelaciones()
    if (fecha_desde or "").strip() and (fecha_hasta or "").strip():
        orders = _get_orders_for_local_date_range(local, fecha_desde.strip()[:10], fecha_hasta.strip()[:10])
    else:
        date_str = (fecha or "").strip()[:10] or _get_today_colombia()
        orders = _get_orders_for_local_date(local, date_str)
    codigos_marcados = {(item.get("codigo") or "").strip() for item in apelaciones.get("items", [])}
    pendientes = []
    for o in orders:
        cod = (o.get("Codigo integracion") or "").strip()
        if cod not in codigos_marcados:
            continue
        ap = _get_apelacion_by_codigo(cod)
        if not ap or ap.get("monto_devuelto") is not None:
            continue  # ya apelada
        if ap.get("descuento_confirmado"):
            continue
        if ap.get("sede_decidio_no_apelar"):
            continue  # sede decidió no apelar; va al apartado Descuentos
        if _apelacion_plazo_vencido(ap, dias_para_apelar):
            continue  # plazo para apelar vencido; ya aparece en Descuentos
        o["apelacion_monto_descontado"] = ap.get("monto_descontado")
        o["apelacion_canal"] = ap.get("canal")
        fotos = _get_fotos_for_order(o)
        o["fotos_entrega"] = fotos.get("entrega", [])
        pendientes.append(o)
    pendientes.sort(key=lambda x: (x.get("Fecha") or x.get("fecha") or ""), reverse=True)
    return {"orders": pendientes}


@app.post("/api/apelaciones/apelar")
async def api_apelar(
    codigo: str = Query(...),
    monto_devuelto: float = Query(...),
    fecha_estimada_devolucion: str = Query("", description="YYYY-MM-DD cuándo nos devolverán el dinero"),
    files: list[UploadFile] = File(default=[]),
):
    """User: sube foto de la respuesta del canal, monto que devolverán y fecha estimada de devolución."""
    cod = (codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no marcada para apelación")
    if ap.get("monto_devuelto") is not None:
        raise HTTPException(status_code=400, detail="Esta orden ya fue apelada")
    data = _read_apelaciones()
    fecha_est = (fecha_estimada_devolucion or "").strip()[:10] or None
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() == cod:
            item["monto_devuelto"] = float(monto_devuelto)
            item["fecha_estimada_devolucion"] = fecha_est
            item["fecha_apelado"] = _now().isoformat()
            break
    _write_apelaciones(data)
    # Guardar fotos de la resolución del canal en apelacion/{canal}/
    _ap_local = (ap.get("local") or "").strip()
    _ap_canal = (ap.get("canal") or "").strip()
    if files:
        safe_canal = _sanitize_path(_ap_canal or "sin_canal") or "sin_canal"
        base = UPLOADS_DIR / _sanitize_codigo(cod) / "apelacion" / safe_canal
        base.mkdir(parents=True, exist_ok=True)
        for f in files:
            if f.filename:
                safe_name = _sanitize_path(f.filename) or "file"
                content = await f.read()
                (base / safe_name).write_bytes(content)
    # Notificar al admin que la sede respondió la apelación
    if _ap_local:
        _create_notificacion(
            local=ADMIN_NOTIF_LOCAL, tipo="apelacion_sede",
            titulo=f"Sede {_ap_local} apelando",
            mensaje=f"{_ap_local} respondió la apelación del pedido #{cod} ({_ap_canal}): van a devolver {_fmt_monto_notif(monto_devuelto)}.",
            route_name="estado-apelaciones",
            extra={"codigo": cod, "local": _ap_local, "canal": _ap_canal, "monto": monto_devuelto},
        )
    return {"ok": True}


class NoApelarBody(BaseModel):
    codigo: str


@app.post("/api/apelaciones/no-apelar")
def api_no_apelar(body: NoApelarBody):
    """La sede decide no apelar: no se descuenta de una, pasa al apartado Descuentos con etiqueta 'La sede decidió no apelar'."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no encontrada en apelaciones")
    if ap.get("monto_devuelto") is not None:
        raise HTTPException(status_code=400, detail="Esta orden ya fue apelada")
    data = _read_apelaciones()
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() == cod:
            item["sede_decidio_no_apelar"] = True
            break
    _write_apelaciones(data)
    return {"ok": True}


@app.get("/api/apelaciones/reembolsos-pendientes")
def api_reembolsos_pendientes(
    local: str = Query("", description="Filtrar por local"),
    fecha_desde: str = Query("", description="YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="YYYY-MM-DD"),
):
    """Órdenes apeladas (con monto_devuelto) que aún no están totalmente reembolsadas (reembolso puede ser incremental)."""
    apelaciones = _read_apelaciones()
    items = [
        i for i in apelaciones.get("items", [])
        if i.get("monto_devuelto") is not None
        and _total_reembolsado(i) < float(i.get("monto_devuelto") or 0)
    ]
    if local:
        items = [i for i in items if (i.get("local") or "").strip() == local.strip()]
    if fecha_desde:
        items = [i for i in items if (i.get("fecha") or "") >= fecha_desde]
    if fecha_hasta:
        items = [i for i in items if (i.get("fecha") or "") <= fecha_hasta]
    for i in items:
        i["total_reembolsado"] = round(_total_reembolsado(i), 2)
        i["reembolsos"] = i.get("reembolsos") if isinstance(i.get("reembolsos"), list) else []
    items.sort(key=lambda x: (x.get("fecha") or ""), reverse=True)
    return {"items": items}


class ReembolsarBody(BaseModel):
    codigo: str
    mismo_valor: bool = True  # True = nos devolvieron monto_devuelto; False = valor diferente
    monto_reembolsado: float | None = None  # Solo si mismo_valor=False
    fecha_reembolso: str = ""  # YYYY-MM-DD


@app.post("/api/apelaciones/reembolsar")
def api_reembolsar(body: ReembolsarBody):
    """Registra un reembolso (incremental). El canal puede devolver en varios pagos."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no encontrada en apelaciones")
    if ap.get("monto_devuelto") is None:
        raise HTTPException(status_code=400, detail="La orden no tiene monto_devuelto")
    monto_devuelto = float(ap.get("monto_devuelto", 0))
    if body.mismo_valor:
        ya_reembolsado = _total_reembolsado(ap)
        monto = max(0, monto_devuelto - ya_reembolsado)
    else:
        monto = float(body.monto_reembolsado or 0)
    if monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")
    fecha = (body.fecha_reembolso or "").strip()[:10] or _now().strftime("%Y-%m-%d")
    data = _read_apelaciones()
    _local_reemb = ""
    _canal_reemb = ""
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        reembolsos = item.get("reembolsos")
        if not isinstance(reembolsos, list):
            reembolsos = []
            if item.get("reembolsado") and item.get("monto_reembolsado") is not None:
                reembolsos.append({
                    "monto": float(item.get("monto_reembolsado") or 0),
                    "fecha": (item.get("fecha_reembolso") or "")[:10] or "",
                })
            item["reembolsos"] = reembolsos
        reembolsos.append({"monto": monto, "fecha": fecha})
        total = sum(float(r.get("monto") or 0) for r in reembolsos if isinstance(r, dict))
        monto_devuelto = float(item.get("monto_devuelto") or 0)
        item["reembolsado"] = total >= monto_devuelto
        item["fecha_reembolso"] = fecha
        item["monto_reembolsado"] = total
        _local_reemb = (item.get("local") or "").strip()
        _canal_reemb = (item.get("canal") or "").strip()
        break
    _write_apelaciones(data)
    if _local_reemb:
        _create_notificacion(
            local=_local_reemb, tipo="reembolso_canal",
            titulo="Reembolso del canal",
            mensaje=f"El canal reembolsó {_fmt_monto_notif(monto)} para el pedido #{cod} ({_canal_reemb}).",
            route_name="mis-apelaciones",
            extra={"codigo": cod, "canal": _canal_reemb, "monto": monto},
        )
    return {"ok": True}


def _calcular_perdida(item: dict) -> float:
    """Pérdida a descontar a la sede = monto_descontado - total reembolsado por el canal (incremental o legacy)."""
    return _calcular_perdida_antes_descuento(item)


# ── Configuración global de la app ────────────────────────────────────────────

_APP_CONFIG_KEY = "app_configuracion"
_APP_CONFIG_DEFAULTS: dict = {"dias_para_apelar": 5}


def _get_app_config() -> dict:
    """Lee la configuración global de la app desde preferencias.json."""
    data = _read_json(PREFERENCIAS_JSON, {})
    if not isinstance(data, dict):
        data = {}
    config = data.get(_APP_CONFIG_KEY)
    if not isinstance(config, dict):
        config = {}
    return {**_APP_CONFIG_DEFAULTS, **config}


def _apelacion_plazo_vencido(item: dict, dias_para_apelar: int) -> bool:
    """True si el plazo para apelar ha vencido y la sede no ha respondido.

    Condición: han pasado >= dias_para_apelar días desde fecha_marcado
    y la sede ni apeló (monto_devuelto is None) ni decidió no apelar.
    """
    if item.get("monto_devuelto") is not None:
        return False
    if item.get("sede_decidio_no_apelar"):
        return False
    fecha_marcado = (item.get("fecha_marcado") or "").strip()
    if not fecha_marcado:
        return False
    try:
        dt_marcado = datetime.fromisoformat(fecha_marcado.replace("Z", "+00:00"))
        if dt_marcado.tzinfo is None:
            dt_marcado = dt_marcado.replace(tzinfo=timezone.utc)
        dias_transcurridos = (datetime.now(timezone.utc) - dt_marcado).days
        return dias_transcurridos >= dias_para_apelar
    except Exception:
        return False


@app.get("/api/apelaciones/estado-admin")
def api_apelaciones_estado_admin(
    local: str = Query("", description="Filtrar por sede"),
    fecha_desde: str = Query("", description="YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="YYYY-MM-DD"),
):
    """Admin: estado de todas las apelaciones (pendiente apelar, apelada, reembolsada, descuento confirmado)."""
    apelaciones = _read_apelaciones()
    items = []
    for item in apelaciones.get("items", []):
        if local and (item.get("local") or "").strip() != local.strip():
            continue
        if fecha_desde and (item.get("fecha") or "") < fecha_desde:
            continue
        if fecha_hasta and (item.get("fecha") or "") > fecha_hasta:
            continue
        out = dict(item)
        perdida = _calcular_perdida(item)
        total_reemb = _total_reembolsado(item)
        total_descu = _total_descuentos_sede(item)
        out["perdida"] = round(perdida, 2)
        out["total_reembolsado"] = round(total_reemb, 2)
        out["total_descuentos_sede"] = round(total_descu, 2)
        out["perdida_restante"] = round(max(0, perdida - total_descu), 2)
        out["no_reconocido_canal"] = round(
            max(0.0, float(item.get("monto_descontado") or 0) - float(item.get("monto_devuelto") or 0)), 2
        )
        # Lista de estados: reembolso, descuento, y siempre mostrar "La sede decidió no apelar" si aplica
        estados = []
        monto_dev = float(item.get("monto_devuelto") or 0)
        if monto_dev > 0 and total_reemb >= monto_dev:
            estados.append("reembolsada")
        elif monto_dev > 0:
            estados.append("apelada")
        if total_descu >= perdida and perdida > 0:
            estados.append("descuento_confirmado")
        if item.get("sede_decidio_no_apelar"):
            estados.append("sede_decidio_no_apelar")
        # (estados ya puede tener reembolsada o apelada según monto_devuelto y total_reembolsado)
        if not estados:
            if item.get("monto_devuelto") is not None:
                estados = ["apelada"]
            else:
                estados = ["pendiente_apelar"]
        out["reembolsos"] = item.get("reembolsos") if isinstance(item.get("reembolsos"), list) else []
        out["descuentos"] = item.get("descuentos") if isinstance(item.get("descuentos"), list) else []
        out["estados"] = estados
        out["estado"] = estados[-1]  # último para compatibilidad y orden por defecto
        items.append(out)
    items.sort(key=lambda x: (x.get("fecha") or ""), reverse=True)
    return {"items": items}


@app.get("/api/apelaciones/descuentos")
def api_apelaciones_descuentos(
    local: str = Query("", description="Filtrar por sede"),
    fecha_desde: str = Query("", description="YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="YYYY-MM-DD"),
    solo_pendientes: bool = Query(True, description="Solo sin cubrir completamente (perdida_restante > 0)"),
    solo_confirmados: bool = Query(False, description="Solo ya descontados en nómina"),
    solo_programados: bool = Query(False, description="Solo los que tienen quincenas programadas pero no ejecutadas"),
    solo_empresa_asume: bool = Query(False, description="Solo los que la empresa asumió (total o parcialmente)"),
):
    """Órdenes con pérdida a descontar a la sede.

    Solo incluye órdenes donde la sede ya puede ser descontada:
    - La sede decidió no apelar, O
    - La sede apelaron (monto_devuelto registrado), O
    - El plazo para apelar venció (días desde fecha_marcado >= config.dias_para_apelar).
    """
    config = _get_app_config()
    dias_para_apelar = int(config.get("dias_para_apelar") or 5)
    apelaciones = _read_apelaciones()
    items = []
    for item in apelaciones.get("items", []):
        perdida = _calcular_perdida(item)
        if perdida <= 0:
            continue
        if local and (item.get("local") or "").strip() != local.strip():
            continue
        if fecha_desde and (item.get("fecha") or "") < fecha_desde:
            continue
        if fecha_hasta and (item.get("fecha") or "") > fecha_hasta:
            continue
        # Solo incluir si la orden ya está lista para descontar
        vencida = _apelacion_plazo_vencido(item, dias_para_apelar)
        puede_descontar = (
            item.get("sede_decidio_no_apelar")
            or item.get("monto_devuelto") is not None
            or vencida
        )
        if not puede_descontar:
            continue
        total_eje = _total_descuentos_sede(item)
        total_prog = _total_descuentos_programados(item)
        monto_empresa = _monto_empresa_asume(item)
        total_cubierto_prog = total_prog + monto_empresa
        total_cubierto_eje = total_eje + monto_empresa
        descuentos_list = item.get("descuentos") if isinstance(item.get("descuentos"), list) else []
        tiene_pendientes_ejecutar = any(
            isinstance(d, dict) and not d.get("ejecutado", True)
            for d in descuentos_list
        )
        if solo_empresa_asume:
            if monto_empresa <= 0:
                continue
        elif solo_confirmados:
            if total_cubierto_eje < perdida:
                continue
        elif solo_programados:
            if not tiene_pendientes_ejecutar:
                continue
        elif solo_pendientes and total_cubierto_prog >= perdida:
            continue
        out = dict(item)
        out["perdida"] = round(perdida, 2)
        out["total_descuentos_sede"] = round(total_eje, 2)
        out["total_programado"] = round(total_prog, 2)
        out["monto_empresa_asume"] = round(monto_empresa, 2)
        out["empresa_asume"] = monto_empresa > 0
        out["fecha_empresa_asume"] = item.get("fecha_empresa_asume", "")
        out["perdida_restante"] = round(max(0, perdida - total_cubierto_prog), 2)
        out["perdida_restante_ejecutar"] = round(max(0, perdida - total_cubierto_eje), 2)
        out["descuentos"] = descuentos_list
        out["apelacion_vencida"] = vencida
        items.append(out)
    items.sort(key=lambda x: (x.get("fecha") or ""), reverse=True)
    return {"items": items}


class ConfirmarDescuentoBody(BaseModel):
    codigo: str
    monto: float = 0  # Monto descontado en esta quincena (incremental)
    quincena: str = ""  # Ej: "2026-02-1" (1ra quincena feb), "2026-02-2" (2da quincena)
    fecha: str = ""  # Fecha del descuento en formato YYYY-MM-DD (opcional, usa fecha actual si no se envía)


@app.post("/api/apelaciones/confirmar-descuento")
def api_confirmar_descuento(body: ConfirmarDescuentoBody):
    """Registra un descuento ejecutado a la sede (incremental por quincena). Backward-compat: marca ejecutado=True."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no encontrada en apelaciones")
    perdida = _calcular_perdida(ap)
    if perdida <= 0:
        raise HTTPException(status_code=400, detail="Esta orden no tiene pérdida a descontar")
    monto = float(body.monto or 0)
    if monto <= 0:
        raise HTTPException(status_code=400, detail="Indica el monto descontado en esta quincena")
    quincena = (body.quincena or "").strip() or (_now().strftime("%Y-%m") + "-1")
    fecha = (body.fecha or "").strip() or _now().strftime("%Y-%m-%d")
    data = _read_apelaciones()
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        descuentos = item.get("descuentos")
        if not isinstance(descuentos, list):
            descuentos = []
            if item.get("descuento_confirmado") and item.get("fecha_descuento_confirmado"):
                resto = max(0, perdida - monto)
                if resto > 0:
                    descuentos.append({
                        "id": str(uuid_mod.uuid4()),
                        "monto": resto,
                        "quincena": "",
                        "fecha": (item.get("fecha_descuento_confirmado") or "")[:10],
                        "ejecutado": True,
                    })
            item["descuentos"] = descuentos
        descuentos.append({"id": str(uuid_mod.uuid4()), "monto": monto, "quincena": quincena, "fecha": fecha, "ejecutado": True})
        total_eje = sum(float(d.get("monto") or 0) for d in descuentos if isinstance(d, dict) and d.get("ejecutado", True))
        item["descuento_confirmado"] = total_eje >= perdida
        item["fecha_descuento_confirmado"] = fecha
        break
    _write_apelaciones(data)
    return {"ok": True}


class ProgramarDescuentoBody(BaseModel):
    codigo: str
    monto: float = 0
    quincena: str = ""
    fecha: str = ""


@app.post("/api/apelaciones/programar-descuento")
def api_programar_descuento(body: ProgramarDescuentoBody):
    """Programa un descuento futuro a la sede (queda pendiente de ejecutar hasta marcarlo como hecho efectivo)."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no encontrada en apelaciones")
    perdida = _calcular_perdida(ap)
    if perdida <= 0:
        raise HTTPException(status_code=400, detail="Esta orden no tiene pérdida a descontar")
    monto = float(body.monto or 0)
    if monto <= 0:
        raise HTTPException(status_code=400, detail="Indica el monto a programar")
    quincena = (body.quincena or "").strip() or (_now().strftime("%Y-%m") + "-1")
    fecha = (body.fecha or "").strip() or _now().strftime("%Y-%m-%d")
    data = _read_apelaciones()
    _local_prog = ""
    _canal_prog = ""
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        descuentos = item.get("descuentos")
        if not isinstance(descuentos, list):
            descuentos = []
            item["descuentos"] = descuentos
        descuentos.append({"id": str(uuid_mod.uuid4()), "monto": monto, "quincena": quincena, "fecha": fecha, "ejecutado": False})
        _local_prog = (item.get("local") or "").strip()
        _canal_prog = (item.get("canal") or "").strip()
        break
    _write_apelaciones(data)
    if _local_prog:
        _create_notificacion(
            local=_local_prog, tipo="descuento_programado",
            titulo="Descuento programado",
            mensaje=f"Se programó un descuento de {_fmt_monto_notif(monto)} para la quincena {quincena} (pedido #{cod}).",
            route_name="mis-descuentos",
            extra={"codigo": cod, "canal": _canal_prog, "monto": monto, "quincena": quincena},
        )
    return {"ok": True}


class EjecutarDescuentoBody(BaseModel):
    codigo: str
    descuento_id: str


@app.post("/api/apelaciones/ejecutar-descuento")
def api_ejecutar_descuento(body: EjecutarDescuentoBody):
    """Marca un descuento programado como ejecutado (hecho efectivo en nómina)."""
    cod = (body.codigo or "").strip()
    did = (body.descuento_id or "").strip()
    if not cod or not did:
        raise HTTPException(status_code=400, detail="codigo y descuento_id requeridos")
    data = _read_apelaciones()
    found = False
    _local_eje = ""
    _canal_eje = ""
    _monto_eje = 0.0
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        descuentos = item.get("descuentos", [])
        for d in descuentos:
            if isinstance(d, dict) and d.get("id") == did:
                d["ejecutado"] = True
                found = True
                break
        if found:
            perdida = _calcular_perdida(item)
            total_eje = sum(float(d.get("monto") or 0) for d in descuentos if isinstance(d, dict) and d.get("ejecutado", True))
            item["descuento_confirmado"] = total_eje >= perdida
            item["fecha_descuento_confirmado"] = _now().strftime("%Y-%m-%d")
            _local_eje = (item.get("local") or "").strip()
            _canal_eje = (item.get("canal") or "").strip()
            _monto_eje = next((float(d.get("monto") or 0) for d in descuentos if isinstance(d, dict) and d.get("id") == did), 0)
            break
    if not found:
        raise HTTPException(status_code=404, detail="Descuento programado no encontrado")
    _write_apelaciones(data)
    if _local_eje:
        _create_notificacion(
            local=_local_eje, tipo="descuento_ejecutado",
            titulo="Descuento aplicado en nómina",
            mensaje=f"Se descontó {_fmt_monto_notif(_monto_eje)} de tu nómina por el pedido #{cod} ({_canal_eje}).",
            route_name="mis-descuentos",
            extra={"codigo": cod, "canal": _canal_eje, "monto": _monto_eje},
        )
    return {"ok": True}


class AsumirEmpresaBody(BaseModel):
    codigo: str
    monto: float = 0  # 0 = asumir el total restante


@app.post("/api/apelaciones/asumir-empresa")
def api_asumir_empresa(body: AsumirEmpresaBody):
    """La empresa asume la pérdida (o parte de ella): no se descuenta a la sede ni a nadie."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    ap = _get_apelacion_by_codigo(cod)
    if not ap:
        raise HTTPException(status_code=404, detail="Orden no encontrada en apelaciones")
    perdida = _calcular_perdida(ap)
    if perdida <= 0:
        raise HTTPException(status_code=400, detail="Esta orden no tiene pérdida")
    data = _read_apelaciones()
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        total_prog = _total_descuentos_programados(item)
        monto_ya = _monto_empresa_asume(item)
        restante = max(0.0, perdida - total_prog - monto_ya)
        monto = float(body.monto or 0)
        if monto <= 0:
            monto = restante  # asumir todo lo restante
        if monto <= 0:
            raise HTTPException(status_code=400, detail="No hay monto restante por asumir")
        nuevo_total = round(monto_ya + monto, 2)
        item["monto_empresa_asume"] = nuevo_total
        item["empresa_asume"] = True
        item["fecha_empresa_asume"] = _now().strftime("%Y-%m-%d")
        # Si la pérdida queda totalmente cubierta (sede + empresa), marcar resuelto
        total_eje = _total_descuentos_sede(item)
        if total_eje + nuevo_total >= perdida:
            item["descuento_confirmado"] = True
        break
    _write_apelaciones(data)
    return {"ok": True}


@app.post("/api/apelaciones/deshacer-empresa-asume")
def api_deshacer_empresa_asume(body: AsumirEmpresaBody):
    """Deshace la absorción de la empresa para un pedido (para corregir errores)."""
    cod = (body.codigo or "").strip()
    if not cod:
        raise HTTPException(status_code=400, detail="codigo requerido")
    data = _read_apelaciones()
    for item in data.get("items", []):
        if (item.get("codigo") or "").strip() != cod:
            continue
        item["monto_empresa_asume"] = 0
        item["empresa_asume"] = False
        item["fecha_empresa_asume"] = ""
        perdida = _calcular_perdida(item)
        total_eje = _total_descuentos_sede(item)
        item["descuento_confirmado"] = total_eje >= perdida
        break
    _write_apelaciones(data)
    return {"ok": True}


@app.get("/api/informes")
def api_informes(
    fecha_desde: str = Query(..., description="YYYY-MM-DD"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD"),
    local: list[str] = Query(default=[], description="Filtrar por sede(s)"),
):
    """Métricas y series para la sección Reportes: órdenes, apelaciones, reembolsos, pérdida por día/sede/canal."""
    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    if not desde or not hasta:
        raise HTTPException(status_code=400, detail="fecha_desde y fecha_hasta requeridos (YYYY-MM-DD)")
    from collections import defaultdict

    today_str = datetime.now().strftime("%Y-%m-%d")
    locales_filter: set[str] = {l.strip() for l in local if l.strip()}

    # Órdenes por día, sede y canal (desde cache deliverys)
    ordenes_por_dia: dict[str, int] = defaultdict(int)
    ordenes_por_sede: dict[str, int] = defaultdict(int)
    ordenes_por_canal: dict[str, int] = defaultdict(int)
    locales_data = _locales_list_for_iteration()
    total_ordenes = 0
    for item in locales_data:
        local_id = _locale_id(item) if isinstance(item, dict) else ""
        local_name = _locale_name(item)
        if not local_id:
            continue
        if locales_filter and local_name not in locales_filter:
            continue
        local_dir = DELIVERYS_CACHE_DIR / local_id
        if not local_dir.is_dir():
            continue
        for json_file in sorted(local_dir.glob("*.json")):
            date_str = json_file.stem
            if date_str < desde or date_str > hasta:
                continue
            cached = _read_json(json_file, {})
            data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                order = _delivery_row_to_order(row)
                cod = (order.get("Codigo integracion") or "").strip()
                if not cod or cod == "—":
                    continue
                total_ordenes += 1
                ordenes_por_dia[date_str] += 1
                ordenes_por_sede[local_name] += 1
                canal = (order.get("Canal de delivery") or "").strip() or "—"
                ordenes_por_canal[canal] += 1

    # Apelaciones en rango: totales y por día/sede/canal
    apelaciones = _read_apelaciones()
    items_ap = [
        i for i in apelaciones.get("items", [])
        if (i.get("fecha") or "") >= desde and (i.get("fecha") or "") <= hasta
        and (not locales_filter or (i.get("local") or "").strip() in locales_filter)
    ]

    apelaciones_por_dia: dict[str, int] = defaultdict(int)
    apelaciones_por_sede: dict[str, int] = defaultdict(int)
    apelaciones_por_canal: dict[str, int] = defaultdict(int)
    reembolsos_por_dia: dict[str, int] = defaultdict(int)
    perdida_por_dia: dict[str, float] = defaultdict(float)
    perdida_por_sede: dict[str, float] = defaultdict(float)
    perdida_por_canal: dict[str, float] = defaultdict(float)

    # Counters for new metrics
    total_descontado = 0.0
    total_devuelto = 0.0
    total_reembolsos = 0
    cnt_sin_apelar = 0
    cnt_apeladas = 0
    cnt_sede_no_apelar = 0
    cnt_reembolsadas = 0
    cnt_pendiente_reembolso = 0
    cnt_reembolso_retraso = 0
    monto_pendiente_reembolso = 0.0
    total_descuentos_sede = 0.0
    monto_pendiente_descuento = 0.0
    cnt_descuento_retraso = 0
    total_empresa_asume = 0.0

    # Per-sede detailed counters
    sede_sin_apelar: dict[str, int] = defaultdict(int)
    sede_apeladas: dict[str, int] = defaultdict(int)
    sede_sede_no_apelar: dict[str, int] = defaultdict(int)
    sede_reembolsadas: dict[str, int] = defaultdict(int)
    sede_pendiente_reembolso: dict[str, int] = defaultdict(int)
    sede_reembolso_retraso: dict[str, int] = defaultdict(int)
    sede_pendiente_descuento: dict[str, float] = defaultdict(float)
    sede_descuento_retraso: dict[str, int] = defaultdict(int)
    sede_empresa_asume: dict[str, float] = defaultdict(float)

    for i in items_ap:
        f = (i.get("fecha") or "")[:10]
        loc = (i.get("local") or "").strip() or "—"
        can = (i.get("canal") or "").strip() or "—"
        apelaciones_por_dia[f] += 1
        apelaciones_por_sede[loc] += 1
        apelaciones_por_canal[can] += 1
        perd = _calcular_perdida(i)
        perdida_por_dia[f] += perd
        perdida_por_sede[loc] += perd
        perdida_por_canal[can] += perd

        monto_desc = float(i.get("monto_descontado") or 0)
        total_descontado += monto_desc

        monto_dev = float(i.get("monto_devuelto") or 0) if i.get("monto_devuelto") is not None else 0.0
        if i.get("monto_devuelto") is not None:
            total_devuelto += monto_dev

        total_reemb = _total_reembolsado(i)

        # Estado apelación
        is_sede_no_apelar = bool(i.get("sede_decidio_no_apelar"))
        is_apelada = bool(i.get("fecha_apelado")) and not is_sede_no_apelar
        is_sin_apelar = not is_apelada and not is_sede_no_apelar

        if is_sin_apelar:
            cnt_sin_apelar += 1
            sede_sin_apelar[loc] += 1
        elif is_apelada:
            cnt_apeladas += 1
            sede_apeladas[loc] += 1
        if is_sede_no_apelar:
            cnt_sede_no_apelar += 1
            sede_sede_no_apelar[loc] += 1

        # Reembolso
        is_fully_reembolsada = monto_dev > 0 and total_reemb >= monto_dev
        is_pendiente_reembolso = monto_dev > 0 and total_reemb < monto_dev
        fecha_est = (i.get("fecha_estimada_devolucion") or "")[:10]
        is_reembolso_retraso = is_pendiente_reembolso and bool(fecha_est) and fecha_est < today_str

        if is_fully_reembolsada:
            total_reembolsos += 1
            cnt_reembolsadas += 1
            reembolsos_por_dia[f] += 1
            sede_reembolsadas[loc] += 1
        if is_pendiente_reembolso:
            cnt_pendiente_reembolso += 1
            monto_pendiente_reembolso += max(0.0, monto_dev - total_reemb)
            sede_pendiente_reembolso[loc] += 1
        if is_reembolso_retraso:
            cnt_reembolso_retraso += 1
            sede_reembolso_retraso[loc] += 1

        # Descuento a sede
        empresa_asume = float(i.get("monto_empresa_asume") or 0)
        total_empresa_asume += empresa_asume
        sede_empresa_asume[loc] += empresa_asume

        descu_ejecutados = _total_descuentos_sede(i)
        total_descuentos_sede += descu_ejecutados

        perd_restante = float(i.get("perdida_restante") or perd)
        if perd_restante > 0:
            monto_pendiente_descuento += perd_restante
            sede_pendiente_descuento[loc] += perd_restante

        # Descuento con retraso: quincenas programadas no ejecutadas con fecha pasada
        descuentos_list = i.get("descuentos") or []
        has_retraso_desc = any(
            not d.get("ejecutado", True) and (d.get("fecha") or "")[:10] < today_str
            for d in descuentos_list if isinstance(d, dict)
        )
        if has_retraso_desc:
            cnt_descuento_retraso += 1
            sede_descuento_retraso[loc] += 1

    total_perdido = round(total_descontado - total_devuelto, 2)
    total_descontado = round(total_descontado, 2)
    total_devuelto = round(total_devuelto, 2)
    ordenes_con_apelacion = len(items_ap)
    pct_con_problemas = round(ordenes_con_apelacion / total_ordenes * 100, 1) if total_ordenes else 0.0

    # Fechas en rango para por_dia
    start = datetime.strptime(desde, "%Y-%m-%d").date()
    end = datetime.strptime(hasta, "%Y-%m-%d").date()
    por_dia = []
    d = start
    while d <= end:
        key = d.strftime("%Y-%m-%d")
        por_dia.append({
            "fecha": key,
            "ordenes": ordenes_por_dia.get(key, 0),
            "apelaciones": apelaciones_por_dia.get(key, 0),
            "reembolsos": reembolsos_por_dia.get(key, 0),
            "perdida": round(perdida_por_dia.get(key, 0), 2),
        })
        d += timedelta(days=1)

    all_sedes = sorted(set(ordenes_por_sede) | set(apelaciones_por_sede))
    por_sede = []
    for loc in all_sedes:
        ords = ordenes_por_sede.get(loc, 0)
        aps = apelaciones_por_sede.get(loc, 0)
        por_sede.append({
            "local": loc,
            "ordenes": ords,
            "apelaciones": aps,
            "pct_problemas": round(aps / ords * 100, 1) if ords else 0.0,
            "sin_apelar": sede_sin_apelar.get(loc, 0),
            "apeladas": sede_apeladas.get(loc, 0),
            "sede_no_apelar": sede_sede_no_apelar.get(loc, 0),
            "reembolsadas": sede_reembolsadas.get(loc, 0),
            "pendiente_reembolso": sede_pendiente_reembolso.get(loc, 0),
            "reembolso_retraso": sede_reembolso_retraso.get(loc, 0),
            "pendiente_descuento": round(sede_pendiente_descuento.get(loc, 0), 2),
            "descuento_retraso": sede_descuento_retraso.get(loc, 0),
            "empresa_asume": round(sede_empresa_asume.get(loc, 0), 2),
            "perdida": round(perdida_por_sede.get(loc, 0), 2),
        })

    por_canal = [
        {
            "canal": can,
            "ordenes": ordenes_por_canal.get(can, 0),
            "apelaciones": apelaciones_por_canal.get(can, 0),
            "perdida": round(perdida_por_canal.get(can, 0), 2),
        }
        for can in sorted(set(ordenes_por_canal) | set(apelaciones_por_canal))
    ]

    return {
        "resumen": {
            "total_ordenes": total_ordenes,
            "ordenes_con_apelacion": ordenes_con_apelacion,
            "ordenes_sin_apelacion": total_ordenes - ordenes_con_apelacion,
            "pct_con_problemas": pct_con_problemas,
            # Estado apelación
            "sin_apelar": cnt_sin_apelar,
            "apeladas": cnt_apeladas,
            "sede_no_apelar": cnt_sede_no_apelar,
            # Reembolso del canal
            "total_descontado_canal": total_descontado,
            "total_devuelto": total_devuelto,
            "total_reembolsos": cnt_reembolsadas,
            "pendiente_reembolso": cnt_pendiente_reembolso,
            "monto_pendiente_reembolso": round(monto_pendiente_reembolso, 2),
            "reembolso_retraso": cnt_reembolso_retraso,
            # Descuento a sede
            "total_descuentos_sede": round(total_descuentos_sede, 2),
            "monto_pendiente_descuento": round(monto_pendiente_descuento, 2),
            "empresa_asume": round(total_empresa_asume, 2),
            "descuento_retraso": cnt_descuento_retraso,
            # Pérdida neta
            "total_perdida": total_perdido,
        },
        "por_dia": por_dia,
        "por_sede": por_sede,
        "por_canal": por_canal,
    }


@app.get("/api/apelaciones/reporte")
def api_apelaciones_reporte(
    local: str = Query("", description="Filtrar por local"),
    fecha_desde: str = Query("", description="YYYY-MM-DD"),
    fecha_hasta: str = Query("", description="YYYY-MM-DD"),
):
    """Reporte: total descontado, devuelto y perdido."""
    apelaciones = _read_apelaciones()
    items = apelaciones.get("items", [])
    if local:
        items = [i for i in items if (i.get("local") or "").strip() == local.strip()]
    if fecha_desde:
        items = [i for i in items if (i.get("fecha") or "") >= fecha_desde]
    if fecha_hasta:
        items = [i for i in items if (i.get("fecha") or "") <= fecha_hasta]
    total_descontado = sum(float(item.get("monto_descontado") or 0) for item in items)
    total_devuelto = sum(float(item.get("monto_devuelto") or 0) for item in items if item.get("monto_devuelto") is not None)
    total_perdido = total_descontado - total_devuelto
    return {
        "total_descontado": round(total_descontado, 2),
        "total_devuelto": round(total_devuelto, 2),
        "total_perdido": round(total_perdido, 2),
        "items": items,
    }


def _build_reporte_rows(
    local: str = "",
    locales_filter: list[str] | None = None,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    filter_val: str = "",
) -> list[dict]:
    """Construye la lista completa de filas del reporte maestro (sin paginar)."""
    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    # Normalize multi-sede filter: combine legacy single `local` with new list
    _locales_set: set[str] = set()
    if locales_filter:
        _locales_set = {l.strip() for l in locales_filter if l.strip()}
    elif local and local.strip():
        _locales_set = {local.strip()}
    locales_data = _locales_list_for_iteration()
    apelaciones = _read_apelaciones()
    apelaciones_by_cod = {(a.get("codigo") or "").strip(): a for a in apelaciones.get("items", []) if (a.get("codigo") or "").strip()}
    rows_list: list[dict] = []
    for item in locales_data:
        local_id = _locale_id(item) if isinstance(item, dict) else ""
        local_name = _locale_name(item)
        if not local_id:
            continue
        if _locales_set and local_name not in _locales_set:
            continue
        local_dir = DELIVERYS_CACHE_DIR / local_id
        if not local_dir.is_dir():
            continue
        for json_file in sorted(local_dir.glob("*.json")):
            date_str = json_file.stem
            if date_str < desde or date_str > hasta:
                continue
            cached = _read_json(json_file, {})
            data = cached.get("data") if isinstance(cached.get("data"), list) else []
            for row in data:
                order = _delivery_row_to_order(row)
                cod = (order.get("Codigo integracion") or "").strip()
                if not cod or cod == "—":
                    continue
                ap = apelaciones_by_cod.get(cod)
                fotos = _get_fotos_for_order(order)
                perdida = round(_calcular_perdida(ap), 2) if ap else 0
                total_reemb = _total_reembolsado(ap) if ap else 0
                total_descu = _total_descuentos_sede(ap) if ap else 0
                total_prog = _total_descuentos_programados(ap) if ap else 0
                monto_empresa = _monto_empresa_asume(ap) if ap else 0
                perdida_restante = round(max(0, perdida - total_prog - monto_empresa), 2) if ap else 0
                perdida_restante_ejecutar = round(max(0, perdida - total_descu - monto_empresa), 2) if ap else 0
                no_reconocido_canal = round(
                    max(0.0, float(ap.get("monto_descontado") or 0) - float(ap.get("monto_devuelto") or 0)), 2
                ) if ap else 0
                estados_apel: list[str] = []
                if ap:
                    monto_dev = float(ap.get("monto_devuelto") or 0)
                    if monto_dev > 0 and total_reemb >= monto_dev:
                        estados_apel.append("reembolsada")
                    elif monto_dev > 0:
                        estados_apel.append("apelada")
                    if total_descu >= perdida and perdida > 0:
                        estados_apel.append("descuento_confirmado")
                    if ap.get("sede_decidio_no_apelar"):
                        estados_apel.append("sede_decidio_no_apelar")
                    if not estados_apel:
                        if monto_dev > 0:
                            estados_apel = ["apelada"]
                        else:
                            estados_apel = ["pendiente_apelar"]
                r = {
                    "local": local_name,
                    "fecha": date_str,
                    "codigo": cod,
                    "canal": order.get("Canal de delivery"),
                    "cliente": order.get("Cliente"),
                    "monto_pagado": order.get("Monto pagado"),
                    "hora": order.get("Hora"),
                    "delivery_id": order.get("delivery_id"),
                    "has_entrega_photo": _order_has_entrega_photo_from_order(order),
                    "fotos_entrega": fotos.get("entrega", []),
                    "fotos_apelacion": fotos.get("apelacion", {}),
                    "apelacion": {
                        "monto_descontado": ap.get("monto_descontado"),
                        "monto_devuelto": ap.get("monto_devuelto"),
                        "monto_reembolsado": round(total_reemb, 2),
                        "fecha_apelado": ap.get("fecha_apelado"),
                        "fecha_estimada_devolucion": ap.get("fecha_estimada_devolucion"),
                        "reembolsado": total_reemb >= float(ap.get("monto_devuelto") or 0) if ap.get("monto_devuelto") is not None else ap.get("reembolsado"),
                        "fecha_reembolso": ap.get("fecha_reembolso"),
                        "reembolsos": ap.get("reembolsos") if isinstance(ap.get("reembolsos"), list) else [],
                        "descuento_confirmado": total_descu >= perdida and perdida > 0,
                        "fecha_descuento_confirmado": ap.get("fecha_descuento_confirmado"),
                        "descuentos": ap.get("descuentos") if isinstance(ap.get("descuentos"), list) else [],
                        "total_descuentos_sede": round(total_descu, 2),
                        "total_programado": round(total_prog, 2),
                        "perdida_restante": perdida_restante,
                        "perdida_restante_ejecutar": perdida_restante_ejecutar,
                        "monto_empresa_asume": round(monto_empresa, 2),
                        "empresa_asume": monto_empresa > 0,
                        "fecha_empresa_asume": ap.get("fecha_empresa_asume"),
                        "fecha_marcado": ap.get("fecha_marcado"),
                        "sede_decidio_no_apelar": ap.get("sede_decidio_no_apelar"),
                    } if ap else None,
                    "estados_apelacion": estados_apel,
                    "estado_apelacion": estados_apel[-1] if estados_apel else "",
                    "perdida": perdida,
                    "no_reconocido_canal": no_reconocido_canal,
                }
                rows_list.append(r)
    rows_list.sort(key=lambda x: (x.get("fecha") or "", x.get("local") or "", x.get("codigo") or ""))
    fv = (filter_val or "").strip().lower()
    if fv:
        rows_list = [
            r for r in rows_list
            if fv in (r.get("local") or "").lower()
            or fv in (r.get("codigo") or "").lower()
            or fv in (r.get("canal") or "").lower()
            or fv in (r.get("cliente") or "").lower()
        ]
    return rows_list


@app.get("/api/reporte-maestro")
def api_reporte_maestro(
    local: list[str] = Query(default=[], description="Filtrar por sede(s) (vacío = todas)"),
    fecha_desde: str = Query(..., description="YYYY-MM-DD"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD"),
    first: int = Query(0, ge=0, description="Índice del primer registro (paginación)"),
    rows: int = Query(20, ge=1, le=500, description="Registros por página"),
    filter: str = Query("", description="Filtro global (local, código, canal, cliente)"),
):
    """Admin: reporte maestro con órdenes + apelaciones + fotos. Paginación real (first/rows) y filtro opcional."""
    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    if not desde or not hasta:
        raise HTTPException(status_code=400, detail="fecha_desde y fecha_hasta requeridos (YYYY-MM-DD)")
    rows_list = _build_reporte_rows(locales_filter=local, fecha_desde=desde, fecha_hasta=hasta, filter_val=filter)
    total_records = len(rows_list)
    page = rows_list[first: first + rows]
    return {"rows": page, "totalRecords": total_records}


@app.get("/api/reporte-maestro/excel")
def api_reporte_maestro_excel(
    request: Request,
    local: list[str] = Query(default=[], description="Filtrar por sede(s) (vacío = todas)"),
    fecha_desde: str = Query(..., description="YYYY-MM-DD"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD"),
    filter: str = Query("", description="Filtro global"),
):
    """
    Descarga el reporte maestro como Excel: una hoja por sede + hoja TODO.
    Fotos como vínculos (Foto 1, Foto 2…); si hay más de una foto por grupo
    se agregan filas extra — esa fila solo lleva el código de referencia y las fotos adicionales.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    import io

    desde = (fecha_desde or "").strip()[:10]
    hasta = (fecha_hasta or "").strip()[:10]
    if not desde or not hasta:
        raise HTTPException(status_code=400, detail="fecha_desde y fecha_hasta requeridos (YYYY-MM-DD)")

    all_rows = _build_reporte_rows(locales_filter=local, fecha_desde=desde, fecha_hasta=hasta, filter_val=filter)
    base_url = str(request.base_url).rstrip("/")

    ESTADO_LABELS: dict[str, str] = {
        "pendiente_apelar": "Pend. apelar",
        "apelada": "Apelada",
        "reembolsada": "Reembolsada",
        "descuento_confirmado": "Descontado",
        "sede_decidio_no_apelar": "Sede no apeló",
    }

    MAX_FOTOS = 2

    # Columnas fijas
    HEADERS = [
        "Sede", "Fecha", "Canal", "Código", "Cliente", "Valor pedido",
        "Foto entrega", "Foto apelación",
        "Descontado por canal", "Reconocido por canal", "No reconocido",
        "Estado reembolso", "Total reembolsado", "Reembolsos detalle",
        "Descontado a sede", "Total programado", "Quincenas",
        "Pend. ejecutar", "Empresa asume", "Pérdida empresa",
        "Estado apelación",
        "Fecha marcado", "Fecha apelado", "Fecha est. devolución",
    ]
    # índices 1-based
    COL_SEDE, COL_FECHA, COL_CANAL, COL_COD, COL_CLI, COL_VALOR = 1, 2, 3, 4, 5, 6
    COL_FOTO_E, COL_FOTO_A = 7, 8
    COL_DESC_CANAL, COL_RECON, COL_NO_RECON = 9, 10, 11
    COL_ESTADO_REIMB, COL_TOTAL_REIMB, COL_REIMB_DETALLE = 12, 13, 14
    COL_DESC_SEDE, COL_TOTAL_PROG, COL_QUINCENAS = 15, 16, 17
    COL_PEND_EJEC, COL_EMPRESA, COL_PERD = 18, 19, 20
    COL_ESTADO_AP = 21
    COL_F_MARCADO, COL_F_APELADO, COL_F_EST_DEV = 22, 23, 24
    FOTO_COLS = {COL_FOTO_E, COL_FOTO_A}
    MONEY_COLS = {COL_VALOR, COL_DESC_CANAL, COL_RECON, COL_NO_RECON,
                  COL_TOTAL_REIMB, COL_DESC_SEDE, COL_TOTAL_PROG,
                  COL_PEND_EJEC, COL_EMPRESA, COL_PERD}

    def to_abs(p: str) -> str:
        p = (p or "").strip()
        return p if p.startswith("http") else base_url + "/" + p.lstrip("/")

    def foto_link(url: str, n: int) -> str:
        safe = url.replace('"', "%22")
        return f'=HYPERLINK("{safe}","Foto {n}")'

    def parse_monto(v: object) -> float | None:
        if v is None:
            return None
        try:
            val = float(str(v).replace(",", "").replace("$", "").replace("\xa0", "").strip() or 0)
            return val if val != 0 else None
        except Exception:
            return None

    def _fmt_iso(iso: str | None) -> str:
        """Formatea ISO datetime a string legible."""
        if not iso:
            return ""
        try:
            from datetime import timezone as _tz
            d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            d_local = d.astimezone(_tz.utc)
            return d_local.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(iso)[:16]

    def expand_rows(r: dict) -> list[list]:
        """
        Devuelve una lista de filas Excel para esta orden.
        La primera fila tiene todos los datos; las extras solo llevan el código
        de referencia y las fotos adicionales en su posición correspondiente.
        """
        ap = r.get("apelacion") or {}
        e_urls = [to_abs(p) for p in (r.get("fotos_entrega") or []) if p][:MAX_FOTOS]
        a_urls = [to_abs(p) for urls in (r.get("fotos_apelacion") or {}).values() for p in (urls or []) if p][:MAX_FOTOS]
        n_rows = max(1, len(e_urls), len(a_urls))

        estados = r.get("estados_apelacion") or ([r.get("estado_apelacion")] if r.get("estado_apelacion") else [])
        perdida = r.get("perdida", 0) or 0
        total_descu = float(ap.get("total_descuentos_sede") or 0)
        total_prog = float(ap.get("total_programado") or 0)
        pend_ejec = float(ap.get("perdida_restante_ejecutar") or 0)
        monto_empresa = float(ap.get("monto_empresa_asume") or 0)
        perdida_empresa = float(ap.get("perdida_restante") or 0)
        monto_dev = ap.get("monto_devuelto") or 0
        total_reimb = float(ap.get("monto_reembolsado") or 0)

        if ap.get("reembolsado"):
            estado_reemb = "Reembolsado"
        elif ap.get("monto_devuelto") is not None and float(ap.get("monto_devuelto") or 0) > 0:
            estado_reemb = "Pendiente"
        else:
            estado_reemb = ""

        # Detalle reembolsos como texto
        reembolsos_list = ap.get("reembolsos") if isinstance(ap.get("reembolsos"), list) else []
        reimb_detalle = "; ".join(
            f"${int(float(rb.get('monto') or 0)):,} ({rb.get('fecha', '')[:10]})"
            for rb in reembolsos_list if isinstance(rb, dict)
        ) or None

        # Detalle quincenas como texto
        descuentos_list = ap.get("descuentos") if isinstance(ap.get("descuentos"), list) else []
        quincenas_txt = "; ".join(
            f"{'✓' if d.get('ejecutado', True) else '⏳'} {d.get('quincena') or d.get('fecha', '')[:7]} ${int(float(d.get('monto') or 0)):,}"
            for d in descuentos_list if isinstance(d, dict)
        ) or None

        result = []
        for i in range(n_rows):
            is_first = i == 0
            row: list = [None] * len(HEADERS)

            # Foto de cada grupo en la posición i (o None si no hay)
            row[COL_FOTO_E - 1] = foto_link(e_urls[i], i + 1) if i < len(e_urls) else None
            row[COL_FOTO_A - 1] = foto_link(a_urls[i], i + 1) if i < len(a_urls) else None

            if is_first:
                row[COL_SEDE  - 1] = r.get("local", "")
                row[COL_FECHA - 1] = r.get("fecha", "")
                row[COL_CANAL - 1] = r.get("canal", "")
                row[COL_COD   - 1] = r.get("codigo", "")
                row[COL_CLI   - 1] = r.get("cliente", "")
                row[COL_VALOR - 1] = parse_monto(r.get("monto_pagado"))
                row[COL_DESC_CANAL   - 1] = parse_monto(ap.get("monto_descontado"))
                row[COL_RECON        - 1] = parse_monto(monto_dev)
                row[COL_NO_RECON     - 1] = r.get("no_reconocido_canal") or None
                row[COL_ESTADO_REIMB - 1] = estado_reemb or None
                row[COL_TOTAL_REIMB  - 1] = total_reimb if total_reimb > 0 else None
                row[COL_REIMB_DETALLE - 1] = reimb_detalle
                row[COL_DESC_SEDE    - 1] = total_descu if total_descu > 0 else None
                row[COL_TOTAL_PROG   - 1] = total_prog if total_prog > 0 else None
                row[COL_QUINCENAS    - 1] = quincenas_txt
                row[COL_PEND_EJEC    - 1] = pend_ejec if pend_ejec > 0 else None
                row[COL_EMPRESA      - 1] = monto_empresa if monto_empresa > 0 else None
                row[COL_PERD         - 1] = perdida_empresa if perdida > 0 else None
                row[COL_ESTADO_AP    - 1] = ", ".join(ESTADO_LABELS.get(e, e) for e in estados) if estados else None
                row[COL_F_MARCADO    - 1] = _fmt_iso(ap.get("fecha_marcado")) or None
                row[COL_F_APELADO    - 1] = _fmt_iso(ap.get("fecha_apelado")) or None
                row[COL_F_EST_DEV    - 1] = (ap.get("fecha_estimada_devolucion") or "")[:10] or None
            else:
                # Fila extra: solo código de referencia + fotos
                row[COL_COD - 1] = r.get("codigo", "")

            result.append(row)
        return result

    # Estilos
    header_fill   = PatternFill("solid", fgColor="1E4D8C")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align  = Alignment(horizontal="center", vertical="center")
    top_align     = Alignment(vertical="center")
    link_font     = Font(color="0563C1", underline="single", size=10)
    cont_fill     = PatternFill("solid", fgColor="F0F4FA")   # fondo suave filas extra
    money_fmt     = "#,##0.00"
    thin          = Side(style="thin", color="D0D0D0")
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths    = [24, 12, 16, 26, 22, 14,  16, 16, 16,  18, 18, 16, 16, 32, 16, 16, 40, 16, 16, 16, 24, 20, 20, 20]

    def build_sheet(ws: object, data_rows: list[dict]) -> None:
        total_cols = len(HEADERS)
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = cell_border
        ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 36

        r_idx = 2
        for row_data in data_rows:
            expanded = expand_rows(row_data)
            for row_offset, cells in enumerate(expanded):
                is_cont = row_offset > 0
                for c_idx, val in enumerate(cells, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = cell_border
                    if is_cont:
                        cell.fill = cont_fill
                    if c_idx in FOTO_COLS:
                        if val:
                            cell.font = link_font
                        cell.alignment = center_align
                    elif c_idx in MONEY_COLS and val is not None:
                        cell.number_format = money_fmt
                        cell.alignment = top_align
                    else:
                        cell.alignment = top_align
                r_idx += 1

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()
    wb.remove(wb.active)

    by_sede: dict[str, list[dict]] = defaultdict(list)
    for r in all_rows:
        by_sede[r.get("local", "Sin sede")].append(r)

    for sede_name in sorted(by_sede.keys()):
        safe = sede_name[:31].translate(str.maketrans("/\\*?[]:", "-------"))
        ws = wb.create_sheet(title=safe)
        build_sheet(ws, by_sede[sede_name])

    ws_todo = wb.create_sheet(title="TODO")
    build_sheet(ws_todo, all_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"reporte_maestro_{desde}_{hasta}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/orders/{codigo:path}/fotos")
def api_get_fotos(codigo: str):
    """Fotos de la orden agrupadas: entrega, apelación (por canal), respuestas. Busca en todas las referencias (codigo integración e identificador unico)."""
    order = _find_order_by_codigo(codigo)
    return _get_fotos_for_order(order) if order else _get_fotos_for_codigo(codigo)


@app.post("/api/orders/{codigo:path}/fotos")
async def api_upload_fotos(
    codigo: str,
    group: str = Query(..., description="entrega | apelacion"),
    canal: str = Query("", description="Canal de venta (para group=apelacion)"),
    files: list[UploadFile] = File(default=[]),
):
    """Sube fotos para una orden. group=entrega|apelacion; canal solo para apelacion."""
    if group not in ("entrega", "apelacion"):
        raise HTTPException(status_code=400, detail="group debe ser entrega o apelacion")
    if not files:
        raise HTTPException(status_code=400, detail="Envía al menos un archivo")
    cod = (codigo or "").strip().lstrip("#")
    base = UPLOADS_DIR / _sanitize_codigo(cod)
    if group == "apelacion":
        canal_safe = _sanitize_path(canal) if canal else "general"
        base = base / "apelacion" / canal_safe
    else:
        base = base / group
    base.mkdir(parents=True, exist_ok=True)
    saved = []
    max_file_size = 50 * 1024 * 1024  # 50 MB por archivo (alineado con nginx client_max_body_size)
    for f in files:
        if not f.filename:
            continue
        content = await f.read()
        if len(content) > max_file_size:
            raise HTTPException(
                status_code=413,
                detail=f"Archivo '{f.filename}' demasiado grande. Máximo 50 MB por imagen.",
            )
        safe_name = _sanitize_path(f.filename) or "file"
        path = base / safe_name
        path.write_bytes(content)
        saved.append(safe_name)
    if group == "entrega" and saved:
        order = _find_order_by_codigo(codigo)
        if order:
            did = (order.get("delivery_id") or "").strip()
            if did:
                _unmark_no_entregada(did)
    return {"saved": saved, "group": group, "canal": canal or None}


def _uploads_base_for_codigo(codigo: str) -> Path:
    """Carpeta base en uploads para un código. Sin # para URLs. Fallback a carpeta con # si existía antes."""
    cod = (codigo or "").strip().lstrip("#")
    base = UPLOADS_DIR / _sanitize_codigo(cod)
    if base.exists():
        return base
    if cod.isdigit():
        base_alt = UPLOADS_DIR / _sanitize_codigo("#" + cod)
        if base_alt.exists():
            return base_alt
    return base


def _foto_path(codigo: str, group: str, path_rest: str) -> Path:
    """Ruta física del archivo de foto. path_rest = filename (entrega) o canal/filename (apelacion)."""
    base = _uploads_base_for_codigo(codigo)
    if group == "apelacion":
        return base / "apelacion" / path_rest
    return base / group / path_rest


@app.get("/api/orders/{codigo:path}/fotos/{group}/{path_rest:path}")
def api_serve_foto(codigo: str, group: str, path_rest: str):
    """Sirve un archivo de foto. path_rest = filename (entrega) o canal/filename (apelacion)."""
    path = _foto_path(codigo, group, path_rest)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(path)


@app.delete("/api/orders/{codigo:path}/fotos/{group}/{path_rest:path}")
def api_delete_foto(codigo: str, group: str, path_rest: str):
    """Elimina un archivo de foto de la orden."""
    path = _foto_path(codigo, group, path_rest)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    path.unlink()
    return {"deleted": path_rest}


@app.post("/api/admin/organize-foto-refs")
def api_organize_foto_refs():
    """
    Renombra y organiza carpetas en uploads/: las que están por identificador unico
    se mueven a la carpeta del código de integración (referencia canónica).
    Así las fotos que ya tenían carpeta con otra referencia quedan unificadas.
    """
    result = _organize_foto_refs()
    return result


def _report_status_payload() -> dict:
    """Construye el objeto de estado para el WebSocket (datos desde API deliverys cada 5 min)."""
    state = _deliverys_scheduler_state
    now = _now()
    next_at = state.get("next_run_at")
    status = state.get("status", "waiting")
    last_at = state.get("last_report_at")
    last_error = state.get("last_error")
    last_filas = state.get("last_filas", 0)
    interval = state.get("interval_seconds", 300)

    if next_at is None:
        seconds_until_next = 0
        message = "Iniciando... Próxima consulta deliverys en breve."
    elif status == "calling_deliverys":
        seconds_until_next = interval
        message = "Llamando API deliverys (fecha hoy, todos los locales)..."
    elif status == "deliverys_ready":
        delta = (next_at - now).total_seconds()
        seconds_until_next = max(0, int(delta))
        if last_error:
            message = f"Última consulta deliverys: error ({last_error}). Próxima en {seconds_until_next} s."
        else:
            message = f"Deliverys listos ({last_filas} registros). Próxima consulta en {seconds_until_next} s."
    else:
        delta = (next_at - now).total_seconds()
        seconds_until_next = max(0, int(delta))
        message = f"Próxima consulta deliverys en {seconds_until_next} s."

    return {
        "status": status,
        "message": message,
        "seconds_until_next": seconds_until_next,
        "next_run_at": next_at.isoformat() if next_at and hasattr(next_at, "isoformat") else None,
        "last_report_at": last_at.isoformat() if last_at and hasattr(last_at, "isoformat") else None,
        "last_error": last_error,
        "last_filas": last_filas,
        "interval_seconds": interval,
    }


# ---------------------------------------------------------------------------
# Planillas diarias por sede
# ---------------------------------------------------------------------------

_PLANILLA_ALLOWED_EXTENSIONS = {
    ".xlsx", ".xls", ".csv", ".ods", ".pdf", ".png", ".jpg", ".jpeg", ".webp",
}


def _planilla_dir(local_id: str, fecha: str) -> Path:
    """Directorio donde se guarda la planilla de una sede y fecha."""
    return PLANILLAS_DIR / _sanitize_path(local_id) / _sanitize_path(fecha)


def _planilla_list_files(local_id: str, fecha: str) -> list[dict]:
    """Devuelve lista de archivos de planilla para una sede y fecha."""
    d = _planilla_dir(local_id, fecha)
    if not d.exists():
        return []
    files = []
    for f in sorted(d.iterdir()):
        if f.is_file() and f.suffix.lower() in _PLANILLA_ALLOWED_EXTENSIONS:
            stat = f.stat()
            files.append({"nombre": f.name, "tamanio": stat.st_size, "fecha_subida": stat.st_mtime})
    return files


def _planilla_status(local_id: str, fecha: str) -> dict:
    """Devuelve estado de las planillas: subida o pendiente, con lista de archivos."""
    archivos = _planilla_list_files(local_id, fecha)
    return {"subida": len(archivos) > 0, "archivos": archivos}


@app.get("/api/planilla/{local_id}/{fecha}/estado")
def api_planilla_estado(local_id: str, fecha: str):
    """Estado de las planillas diarias de una sede (subida o pendiente)."""
    return _planilla_status(local_id, fecha)


@app.post("/api/planilla/{local_id}/{fecha}")
async def api_planilla_upload(
    local_id: str,
    fecha: str,
    file: UploadFile = File(...),
):
    """Sube un archivo de planilla para una sede y fecha (pueden subirse varios)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="El archivo no tiene nombre")
    ext = Path(file.filename).suffix.lower()
    if ext not in _PLANILLA_ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Extensión no permitida: {ext}. Usa {', '.join(sorted(_PLANILLA_ALLOWED_EXTENSIONS))}",
        )
    content = await file.read()
    max_size = 50 * 1024 * 1024  # 50 MB
    if len(content) > max_size:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (máx. 50 MB)")

    d = _planilla_dir(local_id, fecha)
    d.mkdir(parents=True, exist_ok=True)

    safe_name = _sanitize_path(file.filename) or "planilla" + ext
    dest = d / safe_name
    dest.write_bytes(content)
    # Notificar al admin que una sede subió planilla
    _create_notificacion(
        local=ADMIN_NOTIF_LOCAL, tipo="planilla_subida",
        titulo=f"Planilla subida — {local_id}",
        mensaje=f"La sede {local_id} subió la planilla del {fecha}: {safe_name}.",
        route_name="planillas",
        extra={"local_id": local_id, "fecha": fecha, "nombre": safe_name},
    )
    return {"subida": True, "nombre": safe_name}


@app.delete("/api/planilla/{local_id}/{fecha}")
def api_planilla_delete(
    local_id: str,
    fecha: str,
    nombre: str | None = Query(None, description="Nombre del archivo a eliminar; si se omite, elimina todos"),
):
    """Elimina uno o todos los archivos de planilla de una sede para una fecha."""
    d = _planilla_dir(local_id, fecha)
    if nombre:
        safe_name = _sanitize_path(nombre)
        p = d / safe_name
        if not p.exists() or not p.is_file():
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        p.unlink()
        return {"eliminada": True, "nombre": safe_name}
    # Eliminar todos
    files = _planilla_list_files(local_id, fecha)
    if not files:
        raise HTTPException(status_code=404, detail="No hay planillas para esta sede y fecha")
    for fi in files:
        (d / fi["nombre"]).unlink(missing_ok=True)
    return {"eliminada": True, "count": len(files)}


@app.get("/api/planilla/{local_id}/{fecha}/archivo/{nombre}")
def api_planilla_download(local_id: str, fecha: str, nombre: str):
    """Descarga un archivo de planilla específico de una sede para una fecha."""
    safe_name = _sanitize_path(nombre)
    p = _planilla_dir(local_id, fecha) / safe_name
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(p, filename=p.name)


@app.get("/api/planilla/estado-sedes")
def api_planilla_estado_sedes(fecha: str = Query(..., description="YYYY-MM-DD")):
    """
    Estado de planillas para todas las sedes en una fecha.
    Devuelve lista [{local_id, local_name, subida, archivos:[{nombre, tamanio, fecha_subida}]}].
    """
    locales = _locales_list_for_iteration()
    result = []
    for loc in locales:
        lid = str(loc.get("id", ""))
        name = loc.get("name", lid)
        st = _planilla_status(lid, fecha)
        result.append({
            "local_id": lid,
            "local_name": name,
            **st,
        })
    return {"fecha": fecha, "sedes": result}


@app.websocket("/report/ws")
async def report_status_websocket(websocket: WebSocket):
    """
    WebSocket para seguir en tiempo real el programador de reportes.
    Envía cada segundo: estado, mensaje, segundos hasta la próxima consulta,
    cuándo se llamó el reporte y cuándo estará listo/siguiente consulta.
    """
    await websocket.accept()
    _report_ws_clients.append(websocket)
    try:
        while True:
            payload = _report_status_payload()
            await websocket.send_json(payload)
            await asyncio.sleep(1)
    except WebSocketDisconnect:
        pass
    finally:
        if websocket in _report_ws_clients:
            _report_ws_clients.remove(websocket)


# ── Configuración global de la app (endpoints) ────────────────────────────────

@app.get("/api/configuracion")
def api_get_configuracion():
    """Lee la configuración global de la app (dias_para_apelar, etc.)."""
    return _get_app_config()


@app.put("/api/configuracion")
async def api_put_configuracion(request: Request):
    """Guarda (merge) la configuración global de la app."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Body debe ser JSON válido")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Body debe ser un objeto JSON")
    data = _read_json(PREFERENCIAS_JSON, {})
    if not isinstance(data, dict):
        data = {}
    current = data.get(_APP_CONFIG_KEY)
    if not isinstance(current, dict):
        current = {}
    merged = {**_APP_CONFIG_DEFAULTS, **current, **body}
    data[_APP_CONFIG_KEY] = merged
    PREFERENCIAS_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(PREFERENCIAS_JSON, data)
    return {"ok": True, "config": merged}


# ── Preferencias de usuario (visibilidad métricas, orden secciones) ───────────

@app.get("/api/preferencias/{key}")
def api_get_preferencias(key: str):
    """Devuelve las preferencias guardadas para una clave arbitraria."""
    data = _read_json(PREFERENCIAS_JSON, {})
    if not isinstance(data, dict):
        data = {}
    return {"key": key, "data": data.get(key)}


@app.put("/api/preferencias/{key}")
async def api_put_preferencias(key: str, request: Request):
    """Guarda las preferencias para una clave arbitraria."""
    body = await request.json()
    data = _read_json(PREFERENCIAS_JSON, {})
    if not isinstance(data, dict):
        data = {}
    data[key] = body
    PREFERENCIAS_JSON.parent.mkdir(parents=True, exist_ok=True)
    _write_json(PREFERENCIAS_JSON, data)
    return {"ok": True}


# ── Sonido de notificaciones ──────────────────────────────────────────────────

@app.get("/tono.mp3")
def get_tono_mp3():
    """Sirve el archivo de sonido para notificaciones."""
    tono_path = Path(__file__).resolve().parent / "tono.mp3"
    if not tono_path.exists():
        raise HTTPException(status_code=404, detail="tono.mp3 no encontrado")
    return FileResponse(tono_path, media_type="audio/mpeg", headers={"Cache-Control": "public, max-age=86400"})


# ── Notificaciones: REST + WebSocket ─────────────────────────────────────────

@app.get("/api/notificaciones")
def api_get_notificaciones(
    sede: str = Query(..., description="Nombre o ID del local/sede"),
    solo_no_leidas: bool = Query(False),
):
    """Devuelve las notificaciones de una sede, ordenadas de más reciente a más antigua."""
    sede_name = _resolve_local_name(sede)
    data = _read_notificaciones()
    items = [i for i in data.get("items", []) if (i.get("local") or "") == sede_name]
    if solo_no_leidas:
        items = [i for i in items if not i.get("leida")]
    return {
        "items": items,
        "total_no_leidas": sum(1 for i in items if not i.get("leida")),
        "resolved_sede": sede_name,
    }


@app.post("/api/notificaciones/{notif_id}/leer")
def api_marcar_notificacion_leida(notif_id: str):
    data = _read_notificaciones()
    for item in data.get("items", []):
        if item.get("id") == notif_id:
            item["leida"] = True
            break
    _write_notificaciones(data)
    return {"ok": True}


@app.post("/api/notificaciones/leer-todas")
def api_marcar_todas_leidas(sede: str = Query(...)):
    sede_name = _resolve_local_name(sede)
    data = _read_notificaciones()
    for item in data.get("items", []):
        if item.get("local") == sede_name:
            item["leida"] = True
    _write_notificaciones(data)
    return {"ok": True}


@app.delete("/api/notificaciones/{notif_id}")
def api_eliminar_notificacion(notif_id: str):
    data = _read_notificaciones()
    data["items"] = [i for i in data.get("items", []) if i.get("id") != notif_id]
    _write_notificaciones(data)
    return {"ok": True}


@app.websocket("/ws/notificaciones")
async def ws_notificaciones(websocket: WebSocket, sede: str = Query(...)):
    """WebSocket de notificaciones en tiempo real para una sede."""
    await websocket.accept()
    # Resolver ID numérico al nombre real (los links usan IDs, las notificaciones usan nombres)
    sede_name = _resolve_local_name(sede)
    q: asyncio.Queue = asyncio.Queue()
    _notif_queues.setdefault(sede_name, []).append(q)
    try:
        # Enviar no-leídas actuales al conectar
        data = _read_notificaciones()
        pending = [i for i in data.get("items", []) if i.get("local") == sede_name and not i.get("leida")]
        if pending:
            await websocket.send_json({"type": "initial", "items": pending})
        while True:
            try:
                notif = await asyncio.wait_for(q.get(), timeout=25)
                await websocket.send_json({"type": "notificacion", "data": notif})
            except asyncio.TimeoutError:
                await websocket.send_json({"type": "ping"})
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        try:
            _notif_queues.get(sede_name, []).remove(q)
        except ValueError:
            pass
